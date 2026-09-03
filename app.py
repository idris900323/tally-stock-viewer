'''idris' special project'''
from flask import Flask, jsonify, request, render_template, send_file, Response, session, redirect, url_for
import json
import os
import requests
import sys
import xml.etree.ElementTree as ET
import re
import shutil
import subprocess
import threading
import time
import glob
import logging
import psutil
import pandas as pd
from collections import defaultdict
from io import BytesIO
import openpyxl
from datetime import datetime
from functools import wraps
from logging.handlers import RotatingFileHandler
import zipfile
from urllib.parse import quote

from itsdangerous import TimestampSigner, BadSignature, SignatureExpired
from PIL import Image, ImageOps

import database as db
import image_scanner
from api.search import search_bp, set_search_dependencies
from config import Config
from tally.sync import fetch_from_tally_with_retry
from utils.excel_helpers import load_excel_column, load_excel_rows
from utils.normalize import extract_car_base_name, normalize_text, normalize_lookup_key
from utils.product_normalize import extract_type_and_color

app = Flask(__name__)
app.config.from_object(Config)
app.secret_key = app.config["SECRET_KEY"]
app.register_blueprint(search_bp)

BASE_DIR = os.path.dirname(os.path.abspath(__file__))


def _configure_logging():
    log_dir = app.config["LOG_DIR"]
    if not os.path.isabs(log_dir):
        log_dir = os.path.join(BASE_DIR, log_dir)
    os.makedirs(log_dir, exist_ok=True)
    formatter = logging.Formatter("%(asctime)s %(levelname)s [%(name)s] %(message)s")

    root = logging.getLogger()
    root.setLevel(getattr(logging, app.config["LOG_LEVEL"], logging.INFO))
    if not any(isinstance(h, RotatingFileHandler) for h in root.handlers):
        log_file = app.config["LOG_FILE"]
        if not os.path.isabs(log_file):
            log_file = os.path.join(BASE_DIR, log_file)
        file_handler = RotatingFileHandler(log_file, maxBytes=2 * 1024 * 1024, backupCount=5)
        file_handler.setFormatter(formatter)
        root.addHandler(file_handler)


_configure_logging()
logger = logging.getLogger(__name__)

TALLY_HTTP_SESSION = requests.Session()
EXPORT_LOCK = threading.Lock()
FULL_REFRESH_LOCK = threading.Lock()
full_refresh_status = {
    "running": False,
    "stage": None,
    "error": None,
    "timestamp": None,
}
LOGIN_ATTEMPTS = {}
START_TIME = time.time()

PROJECT_ROOT = BASE_DIR
PROJECT_IMAGE_ROOT = os.path.join(BASE_DIR, "data", "S.S IMAGE")
IMAGE_SCAN_ROOT = os.path.abspath(
    os.environ.get("IMAGE_SCAN_ROOT", PROJECT_IMAGE_ROOT)
)

# Share-optimized image variant (see get_share_image()/_build_share_image()
# below) -- only ever used for the bulk "Share Images" flow, never for the
# main design grid, the full-screen modal, or Copy Image (all of which keep
# serving full-resolution originals). 1280px long edge / JPEG quality 80
# roughly matches what WhatsApp itself re-compresses images down to anyway,
# and measured ~64% average size reduction across a real sample of this
# catalog's images (dramatically more for the large-tail raw phone photos
# that dominate total transfer time in a bulk share).
SHARE_IMAGE_MAX_DIMENSION = 1280
SHARE_IMAGE_JPEG_QUALITY = 80
SHARE_IMAGE_CACHE_DIR = os.path.join(BASE_DIR, "data", "share_cache")
PLACEHOLDER_SVG = b"""<svg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 320 320' role='img' aria-label='No image available'>
<rect width='320' height='320' rx='24' fill='#eef2f7'/>
<rect x='54' y='54' width='212' height='212' rx='18' fill='#d9e2ec'/>
<path d='M84 210l46-46 36 36 26-26 44 44H84z' fill='#8da2b8'/>
<circle cx='121' cy='122' r='20' fill='#8da2b8'/>
<text x='160' y='286' text-anchor='middle' font-family='Arial, sans-serif' font-size='20' fill='#52616b'>No image mapped</text>
</svg>"""

# file locations and polling configuration

CAR_FILE = os.path.join(BASE_DIR, "data", "car master list.xls")  # dropdown list
CAR_MASTER_CACHE_JSON = "data/car_master.json"
MAIN_FILE_CANDIDATES = [
    os.path.join(BASE_DIR, "data", "main.xlsx"),
    os.path.join(BASE_DIR, "data", "main.xls"),
    os.path.join(BASE_DIR, "main.xlsx"),
    os.path.join(BASE_DIR, "main.xls"),
]  # parent-child hierarchy (never overwritten)
MAIN_HIERARCHY_CACHE_JSON = "data/main_hierarchy.json"
ITEM_STOCK_FILE = os.path.join(BASE_DIR, "data", "item stock list.xls")  # flat stock item export (quantities only)
ITEM_STOCK_FILE_XLSX = os.path.join(BASE_DIR, "data", "item stock list.xlsx")  # preferred auto-export target
ITEM_STOCK_FILE_AUTO = os.path.join(BASE_DIR, "data", "item stock list.auto.xlsx")  # runtime auto-export target
ITEM_STOCK_CACHE_JSON = os.path.join(BASE_DIR, "data", "item stock list.auto.json")
ITEM_EXPORT_INTERVAL = app.config["TALLY_EXPORT_INTERVAL"]
item_export_timer = None
item_export_enabled = os.environ.get("AUTO_EXPORT_ITEM", "1").strip().lower() not in ("0", "false")
MAX_IMAGE_RESPONSE_LIMIT = max(1, int(app.config["MAX_IMAGE_RESPONSE_LIMIT"]))

# Configuration for automatic polling of Tally
TALLY_URL = app.config["TALLY_URL"]
TALLY_TIMEOUT = max(120, int(app.config["TALLY_TIMEOUT"]))
TALLY_RETRY_ATTEMPTS = app.config["TALLY_RETRY_ATTEMPTS"]
MAX_RETRY_ATTEMPTS = 3
VALID_ROLES = {"admin", "customer"}
PUBLIC_ENDPOINTS = {"login", "logout", "static", "full_refresh_status_route"}

db.init_database()


def _current_role():
    role = session.get("role")
    return role if role in VALID_ROLES else None


def _current_user_id():
    user_id = session.get("user_id")
    try:
        return int(user_id)
    except (TypeError, ValueError):
        return None


def _safe_next_url(raw_next):
    from urllib.parse import urlparse

    if not raw_next or not isinstance(raw_next, str):
        return url_for("home")

    parsed = urlparse(raw_next)
    if parsed.netloc or parsed.scheme:
        return url_for("home")
    if not parsed.path.startswith("/") or parsed.path.startswith("//"):
        return url_for("home")

    return raw_next


def _admin_denied():
    return jsonify({"error": "admin access required"}), 403


def admin_required(view_func):
    @wraps(view_func)
    def wrapper(*args, **kwargs):
        if _current_role() != "admin":
            return _admin_denied()
        return view_func(*args, **kwargs)

    return wrapper


# ============================================================
# SYSTEM PANEL — device-locked remote admin/deployment panel
# ============================================================
# One-time setup:
#   1. Add a line to .env:
#        SYSTEM_ACCESS_TOKEN=<a long random secret you choose>
#      Nothing is auto-generated here — you must set this deliberately.
#      If it's missing, every /admin/system* route returns 403 and the
#      panel is effectively disabled.
#   2. Restart the app so the new .env value is picked up
#      (config.py now calls load_dotenv() at import time).
#   3. From each browser/device you want to pair, while logged in as
#      admin, visit ONCE:
#        /admin/system/authorize-device?token=<that same SYSTEM_ACCESS_TOKEN>
#      This sets a signed, httponly, 400-day cookie on that browser only.
#      Every /admin/system* route requires BOTH a valid admin session AND
#      this cookie — a stolen admin password alone is not enough to reach
#      this panel from an unpaired device.
# ============================================================
SYSTEM_DEVICE_COOKIE_NAME = "system_device_key"
SYSTEM_DEVICE_COOKIE_VALUE = "system-device-authorized"
SYSTEM_DEVICE_COOKIE_MAX_AGE = 400 * 24 * 3600  # 400 days


def _system_panel_configured():
    return bool(str(Config.SYSTEM_ACCESS_TOKEN or "").strip())


def _system_device_signer():
    return TimestampSigner(app.secret_key)


def _is_system_device_authorized():
    cookie_value = request.cookies.get(SYSTEM_DEVICE_COOKIE_NAME)
    if not cookie_value:
        return False
    try:
        unsigned = _system_device_signer().unsign(cookie_value, max_age=SYSTEM_DEVICE_COOKIE_MAX_AGE)
    except (BadSignature, SignatureExpired):
        return False
    return unsigned.decode("utf-8") == SYSTEM_DEVICE_COOKIE_VALUE


def system_device_required(view_func):
    """Stack alongside @admin_required on every /admin/system* route."""
    @wraps(view_func)
    def wrapper(*args, **kwargs):
        if not _system_panel_configured():
            return jsonify({
                "error": "System panel is not configured. Set SYSTEM_ACCESS_TOKEN in .env to enable it."
            }), 403
        if not _is_system_device_authorized():
            return jsonify({
                "error": "This device is not authorized for System Panel access. "
                         "Visit /admin/system/authorize-device?token=YOUR_TOKEN once from this browser to pair it."
            }), 403
        return view_func(*args, **kwargs)

    return wrapper


# ============================================================
# ACCOUNTS PASSWORD GATE — secondary auth layer on Manage Accounts
# ============================================================
# Even a valid admin session cannot view or modify customer accounts
# without also entering a separate password, so a compromised admin
# session alone isn't enough to reach the most sensitive part of the
# admin panel.
#   1. Set ACCOUNTS_ACCESS_PASSWORD in .env. If unset, every accounts
#      route below returns 403 (same "not configured" pattern as the
#      System panel's SYSTEM_ACCESS_TOKEN above) instead of silently
#      allowing access.
#   2. Unlike the System panel's device pairing (a long-lived cookie),
#      this is purely session-based: session["accounts_unlocked"] only
#      lives as long as the login session does (session.clear() on both
#      /login and /logout already wipes it), so it must be re-entered
#      on every new login.
# ============================================================

def _accounts_password_configured():
    return bool(str(Config.ACCOUNTS_ACCESS_PASSWORD or "").strip())


def _accounts_unlocked():
    return bool(session.get("accounts_unlocked"))


def accounts_access_required(view_func=None, *, is_page=False):
    """Stack alongside (after) @admin_required on every accounts route.

    is_page=True is for the /admin/accounts page itself: when locked it
    renders a password interstitial instead of the real page. Every other
    (JSON) accounts route just returns a 403 when locked.
    """
    def decorator(func):
        @wraps(func)
        def wrapper(*args, **kwargs):
            if not _accounts_password_configured():
                return jsonify({
                    "error": "Accounts panel is not configured. Set ACCOUNTS_ACCESS_PASSWORD in .env to enable it."
                }), 403
            if not _accounts_unlocked():
                if is_page:
                    return render_template(
                        "accounts_unlock.html",
                        error=None,
                        next_url=request.path,
                    ), 200
                return jsonify({
                    "error": "Accounts access is locked. Enter the accounts password first."
                }), 403
            return func(*args, **kwargs)

        return wrapper

    if view_func is not None:
        return decorator(view_func)
    return decorator


@app.before_request
def require_login():
    session.permanent = True
    if request.endpoint in PUBLIC_ENDPOINTS or request.endpoint is None:
        return None
    if _current_role() is None:
        if request.method == "GET":
            return render_template(
                "login.html",
                error=None,
                next_url=request.path or url_for("home"),
            ), 401
        return jsonify({"error": "login required"}), 401


@app.after_request
def add_security_headers(response):
    response.headers["X-Content-Type-Options"] = "nosniff"
    response.headers["X-Frame-Options"] = "DENY"
    response.headers["X-XSS-Protection"] = "1; mode=block"
    return response


@app.context_processor
def inject_session_context():
    role = _current_role()
    # Queried fresh on every render (not cached) -- the categories table is
    # tiny, and this is what lets the Category Settings panel's changes show
    # up the moment any page next renders, with zero extra cache-invalidation
    # wiring. design_categories_json feeds the on-thumbnail ribbon and the
    # "Assign Category" picker on the client (see index.html); design_category_choices
    # keeps the same shape existing templates already iterate over.
    live_categories = db.get_all_categories()
    return {
        "current_role": role,
        "current_user_id": _current_user_id(),
        "current_username": session.get("username", ""),
        "is_admin": role == "admin",
        "is_customer": role == "customer",
        "is_viewer": role == "customer",
        "design_category_choices": [c["name"] for c in live_categories],
        "design_categories_json": live_categories,
    }


def _normalize_text(text):
    if text is None:
        return ""
    return normalize_text(str(text)).strip()


def _normalize_lookup_key(text):
    return normalize_lookup_key(text)


def _sanitize_tally_xml(text):
    import re
    text = re.sub(u"[\x00-\x08\x0b\x0c\x0e-\x1f]", "", text)

    def _strip_invalid_ref(match):
        try:
            code = int(match.group(1))
        except ValueError:
            return ""
        if code in (0x9, 0xA, 0xD) or (0x20 <= code <= 0xD7FF) or (0xE000 <= code <= 0xFFFD):
            return match.group(0)
        return ""

    text = re.sub(r"&#(\d+);", _strip_invalid_ref, text)
    return text


def _file_fingerprint(file_path):
    if not file_path or not os.path.exists(file_path):
        return None
    stats = os.stat(file_path)
    return (os.path.abspath(file_path), stats.st_mtime_ns, stats.st_size)


MULTIPLE_TALLY_MESSAGE = "Multiple Tally windows are open. Close the extra Tally and keep only one."


def _check_multiple_tally_instances():
    count = 0
    for process in psutil.process_iter(["name", "exe"]):
        try:
            names = []
            process_name = str(process.info.get("name") or "").strip().lower()
            if process_name:
                names.append(process_name)
                if process_name.endswith(".exe"):
                    names.append(process_name[:-4])

            process_exe = str(process.info.get("exe") or "").strip().lower()
            if process_exe:
                exe_name = os.path.basename(process_exe)
                if exe_name:
                    names.append(exe_name)
                    if exe_name.endswith(".exe"):
                        names.append(exe_name[:-4])

            if "tally.exe" in names or "tally" in names:
                count += 1
        except (psutil.NoSuchProcess, psutil.AccessDenied, psutil.ZombieProcess):
            continue
    return count


# Runtime caches used by hot endpoints.
STOCK_ITEMS_CACHE = []
MAIN_ROWS_CACHE = {"fingerprint": None, "rows": [], "exact_index": {}}
STOCK_QTY_CACHE = {"fingerprint": None, "qty_map": {}}
PRODUCT_CATEGORY_CACHE = {"fingerprint": None, "categories": []}
HIERARCHY_JSON_CACHE = {"fingerprint": None, "payload": []}
PARENT_NAME_SET = set()
DATA_CACHE_LOCK = threading.Lock()

# Per-car completion stats behind the Needs Category / Needs Image Matching
# dashboard queues (see _compute_car_completion_stats() below). Keyed on
# (hierarchy file fingerprint, _QUEUE_STATS_VERSION) rather than the file
# fingerprint alone: total_items comes from the hierarchy file (which already
# invalidates the usual way), but image_linked_count/category_set_count come
# from the mappings/design_categories tables, which have no file of their own
# to fingerprint -- _QUEUE_STATS_VERSION is bumped by hand at every write
# path that can change those counts (see _invalidate_queue_stats_cache()).
QUEUE_STATS_CACHE = {"cache_key": None, "by_car": {}}
_QUEUE_STATS_VERSION = 0

# Cap on how many specific item names the hybrid queue display (Part 3 of
# the queue restructure) shows per car before falling back to "+N more" --
# keeps each car's queue row scannable instead of dumping its whole item
# list inline.
QUEUE_ITEM_NAME_DISPLAY_CAP = 5


def _invalidate_runtime_caches():
    global STOCK_ITEMS_CACHE, MAIN_ROWS_CACHE, STOCK_QTY_CACHE, PRODUCT_CATEGORY_CACHE, HIERARCHY_JSON_CACHE
    STOCK_ITEMS_CACHE = []
    MAIN_ROWS_CACHE = {"fingerprint": None, "rows": [], "exact_index": {}}
    STOCK_QTY_CACHE = {"fingerprint": None, "qty_map": {}}
    PRODUCT_CATEGORY_CACHE = {"fingerprint": None, "categories": []}
    HIERARCHY_JSON_CACHE = {"fingerprint": None, "payload": []}
    _invalidate_queue_stats_cache()


def _invalidate_queue_stats_cache():
    """Call at every write path that can change a stock item's image-mapped
    or category-assigned status (confirm/remove mapping, remove-missing-images,
    assign_category) plus on Full Refresh, so /api/needs_category_queue and
    /api/needs_image_matching_queue never serve stale completion counts."""
    global _QUEUE_STATS_VERSION
    with DATA_CACHE_LOCK:
        _QUEUE_STATS_VERSION += 1


def _load_main_hierarchy_payload():
    """Cached parse of main_hierarchy.json, invalidated by the file's
    fingerprint (mtime + size). Only the JSON parse of the ~15,000-item
    hierarchy is cached here -- quantities are intentionally NOT part of
    this cache and must still be attached fresh by callers via
    _load_training_stock_qty_map(), since stock quantities update far more
    often (roughly every export cycle) than the hierarchy file itself
    (only on Full Refresh); caching quantities here would show stale
    numbers in Training Mode / Bulk Match between refreshes."""
    fingerprint = _file_fingerprint(MAIN_HIERARCHY_CACHE_JSON)
    if fingerprint is not None:
        with DATA_CACHE_LOCK:
            if HIERARCHY_JSON_CACHE["fingerprint"] == fingerprint:
                return HIERARCHY_JSON_CACHE["payload"]

    try:
        with open(MAIN_HIERARCHY_CACHE_JSON, "r", encoding="utf-8") as handle:
            payload = json.load(handle)
        if not isinstance(payload, list):
            payload = []
    except Exception:
        payload = []

    if payload and fingerprint is not None:
        with DATA_CACHE_LOCK:
            HIERARCHY_JSON_CACHE["fingerprint"] = fingerprint
            HIERARCHY_JSON_CACHE["payload"] = payload

    return payload


def _get_product_categories():
    """Group every stock item across all cars by (type, color), cached until
    main_hierarchy.json changes (mirrors MAIN_ROWS_CACHE's invalidation)."""
    fingerprint = _file_fingerprint(MAIN_HIERARCHY_CACHE_JSON)
    with DATA_CACHE_LOCK:
        if fingerprint is not None and PRODUCT_CATEGORY_CACHE["fingerprint"] == fingerprint:
            return PRODUCT_CATEGORY_CACHE["categories"]

    items = _load_training_hierarchy_items()
    counts = defaultdict(int)
    for item in items:
        type_label, colors = extract_type_and_color(item.get("stock_item_name", ""))
        if not type_label:
            continue
        color_key = "-".join(colors)
        counts[(type_label, color_key)] += 1

    categories = [
        {"type": type_label, "color": color_key, "count": count}
        for (type_label, color_key), count in counts.items()
    ]
    categories.sort(key=lambda entry: entry["count"], reverse=True)

    with DATA_CACHE_LOCK:
        PRODUCT_CATEGORY_CACHE["fingerprint"] = fingerprint
        PRODUCT_CATEGORY_CACHE["categories"] = categories
    return categories


def _all_stock_items():
    global STOCK_ITEMS_CACHE
    if STOCK_ITEMS_CACHE:
        return STOCK_ITEMS_CACHE

    items = []
    seen = set()
    for car_model, designs in CAR_DESIGN_MAP.items():
        for design in designs:
            stock_item = design.get("raw") or design.get("design")
            if not stock_item:
                continue
            normalized = _normalize_text(stock_item)
            if normalized in seen:
                continue
            seen.add(normalized)
            items.append({
                "car_model": car_model,
                "stock_item_name": stock_item,
                "qty": design.get("qty", 0),
            })
    STOCK_ITEMS_CACHE = items
    return STOCK_ITEMS_CACHE



# Display-order rank for each category (Part 5 of the design-list grouping
# spec). Categories are now admin-editable (System panel reordering, Part 6),
# so this is built fresh from the live `categories` table on every call
# rather than once at import time from a fixed tuple -- it's a small table,
# and _build_design_payload() below only calls this once per payload (not
# once per item), so there's no hot-path cost to staying live. Uncategorized
# items sort after every real category.
def _category_sort_rank_map():
    categories = db.get_all_categories()
    rank_map = {c["name"]: rank for rank, c in enumerate(categories)}
    return rank_map, len(categories)


def _build_design_payload(designs):
    stock_item_names = []
    seen = set()
    for item in designs or []:
        stock_item_name = item.get("design") or item.get("raw") or "Unknown"
        lookup_key = _normalize_lookup_key(stock_item_name)
        if lookup_key and lookup_key not in seen:
            seen.add(lookup_key)
            stock_item_names.append(stock_item_name)

    mapping_lookup = db.get_mappings_for_stock_items(stock_item_names)
    category_lookup = db.get_categories_for_stock_items(stock_item_names)
    payload = []
    for item in designs or []:
        stock_item_name = item.get("design") or item.get("raw") or "Unknown"
        lookup_key = _normalize_lookup_key(stock_item_name)
        mapping = mapping_lookup.get(lookup_key)
        enriched_item = dict(item)
        if mapping and mapping.get("image_id"):
            enriched_item.update({
                "mapped": True,
                "image_id": mapping["image_id"],
                "thumbnail_url": f"/get_image/{mapping['image_id']}",
                "confidence": mapping.get("confidence", 1.0),
            })
        else:
            enriched_item.update({
                "mapped": False,
                "image_id": None,
                "thumbnail_url": f"/get_stock_image?stock_item={quote(stock_item_name)}",
                "confidence": 0.0,
            })
        enriched_item["fix_url"] = f"/train?stock_item={quote(stock_item_name)}"
        # Assigned via /admin/assign_category (design_categories table) --
        # visible to both admin and customer views, only ever set server-side
        # through the admin-only route.
        enriched_item["category"] = category_lookup.get(lookup_key)
        payload.append(enriched_item)

    # Group by category in Pearl/Pearl Designer/Pearl Deluxe/Saka/Ruby/Napa
    # Deluxe/Napa Designer/uncategorized-last order (Part 5), for both admin
    # and customer views alike since both go through this one function --
    # sorted() is stable, so within each category group items keep whatever
    # order `designs` already handed in (the existing qty-scan/hierarchy
    # order), no secondary sort key needed. Sharing (see copySelectedImages()
    # in index.html, which reads cards in DOM/render order, not click order)
    # automatically follows this same grouping as a result, with no separate
    # client-side reordering required.
    rank_map, uncategorized_rank = _category_sort_rank_map()
    payload.sort(key=lambda item: rank_map.get(item.get("category"), uncategorized_rank))
    return payload


def _compute_car_completion_stats():
    """Per-car completion stats behind the Needs Category / Needs Image
    Matching dashboard queues -- one pass across the whole catalog rather
    than one hierarchy scan + one DB round-trip per car.

    total_items is qty-agnostic (every hierarchy child under the car, same
    scope as /api/get_all_items_for_car / get_stock_items_for_training_from_
    hierarchy()) and counts raw hierarchy entries, not deduplicated by name --
    matching how _build_design_payload() doesn't dedupe its output either, so
    a car's total_items here always equals the length of what that car's
    /api/get_all_items_for_car response would return.

    Cached under (hierarchy file fingerprint, _QUEUE_STATS_VERSION) -- see
    _invalidate_queue_stats_cache().
    """
    hierarchy_fingerprint = _file_fingerprint(MAIN_HIERARCHY_CACHE_JSON)
    cache_key = (hierarchy_fingerprint, _QUEUE_STATS_VERSION)

    with DATA_CACHE_LOCK:
        if QUEUE_STATS_CACHE["cache_key"] == cache_key:
            return QUEUE_STATS_CACHE["by_car"]

    items = _load_training_hierarchy_items()

    stock_names_by_car = {}
    unique_names_by_key = {}
    for item in items:
        car_model = str(item.get("car_model") or "").strip()
        stock_item_name = str(item.get("stock_item_name") or "").strip()
        if not car_model or not stock_item_name:
            continue
        stock_names_by_car.setdefault(car_model, []).append(stock_item_name)
        key = _normalize_lookup_key(stock_item_name)
        if key and key not in unique_names_by_key:
            unique_names_by_key[key] = stock_item_name

    # get_mappings_for_stock_items() scans the whole mappings table once and
    # filters in Python (no SQL IN-clause), so it's safe to call with every
    # distinct stock item name in the catalog in one shot.
    mapping_lookup = db.get_mappings_for_stock_items(list(unique_names_by_key.values()))

    # get_categories_for_stock_items() DOES use a SQL "IN (...)" clause, and
    # a full catalog's worth of distinct item names can run past SQLite's
    # bound-parameter limit in one call -- chunk it the same way
    # /api/search_all_stock_items caps its own result set, well under any
    # realistic SQLITE_MAX_VARIABLE_NUMBER.
    category_lookup = {}
    unique_names = list(unique_names_by_key.values())
    chunk_size = 500
    for start in range(0, len(unique_names), chunk_size):
        category_lookup.update(db.get_categories_for_stock_items(unique_names[start:start + chunk_size]))

    by_car = {}
    for car_model, stock_names in stock_names_by_car.items():
        total_items = len(stock_names)
        image_linked_count = 0
        category_set_count = 0
        fully_done_count = 0
        # Capped, in-scan-order lists of the actual item names still missing
        # a category / image -- the hybrid queue display's "specific item
        # names" (Part 3 of the queue restructure) reads straight off these
        # instead of a second pass or a separate endpoint. Capped at
        # collection time (QUEUE_ITEM_NAME_DISPLAY_CAP) since the queue only
        # ever shows a handful per car; the *count* of remaining items is
        # still the exact image_linked_count/category_set_count math below,
        # not len() of these lists.
        missing_category_items = []
        missing_image_items = []
        for stock_item_name in stock_names:
            key = _normalize_lookup_key(stock_item_name)
            has_image = bool(mapping_lookup.get(key, {}).get("image_id")) if key else False
            has_category = key in category_lookup if key else False
            if has_image:
                image_linked_count += 1
            elif len(missing_image_items) < QUEUE_ITEM_NAME_DISPLAY_CAP:
                missing_image_items.append(stock_item_name)
            if has_category:
                category_set_count += 1
            elif len(missing_category_items) < QUEUE_ITEM_NAME_DISPLAY_CAP:
                missing_category_items.append(stock_item_name)
            if has_image and has_category:
                fully_done_count += 1
        by_car[car_model] = {
            "car": car_model,
            "total_items": total_items,
            "image_linked_count": image_linked_count,
            "category_set_count": category_set_count,
            "fully_done_count": fully_done_count,
            "missing_category_items": missing_category_items,
            "missing_image_items": missing_image_items,
        }

    with DATA_CACHE_LOCK:
        QUEUE_STATS_CACHE["cache_key"] = cache_key
        QUEUE_STATS_CACHE["by_car"] = by_car

    return by_car


def _compute_category_completion_stats():
    """Category-completion stat for the Train Matches dashboard (Part 4 of
    the queue restructure) -- same Linked/Total/Remaining/Complete shape as
    db.get_mapping_stats(), but for categories instead of images. Sums
    _compute_car_completion_stats()'s per-car totals (the exact numbers
    already driving the queue tiers) rather than a separate query, so this
    can never disagree with what the queues show."""
    stats_by_car = _compute_car_completion_stats()
    total_items = sum(stats["total_items"] for stats in stats_by_car.values())
    categorized_items = sum(stats["category_set_count"] for stats in stats_by_car.values())
    percent = round((categorized_items / total_items) * 100, 2) if total_items else 0.0
    return {
        "total_items": total_items,
        "categorized_items": categorized_items,
        "remaining_items": max(total_items - categorized_items, 0),
        "percent_complete": percent,
    }


def _placeholder_response():
    return Response(PLACEHOLDER_SVG, mimetype="image/svg+xml")


def _resolve_car_model_hint(image_record):
    if not image_record:
        return None

    folder_name = image_record.get("car_folder") or ""
    folder_mapping = db.get_folder_car_model(folder_name)
    if folder_mapping:
        return folder_mapping.get("car_model_name")

    folder_norm = _normalize_text(folder_name)
    for car_model in CAR_GROUPS:
        car_norm = _normalize_text(car_model)
        if car_norm and (car_norm in folder_norm or folder_norm in car_norm):
            return car_model

    return None


def _load_training_stock_qty_map():
    json_file = ITEM_STOCK_CACHE_JSON
    if os.path.exists(json_file):
        try:
            with open(json_file, "r", encoding="utf-8") as handle:
                payload = json.load(handle)
            qty_map = {}
            for row in payload if isinstance(payload, list) else []:
                item_name = str(row.get("item_name") or "").strip()
                try:
                    qty = int(row.get("qty") or 0)
                except (TypeError, ValueError):
                    continue
                if item_name and qty > 0:
                    qty_map[_normalize_text(item_name)] = qty
            if qty_map:
                with DATA_CACHE_LOCK:
                    STOCK_QTY_CACHE["fingerprint"] = _file_fingerprint(json_file)
                    STOCK_QTY_CACHE["qty_map"] = qty_map
                return qty_map
        except Exception:
            pass

    stock_file = get_latest_stock_file_path()
    fingerprint = _file_fingerprint(stock_file)
    if fingerprint is None:
        return {}

    with DATA_CACHE_LOCK:
        if STOCK_QTY_CACHE["fingerprint"] == fingerprint:
            return STOCK_QTY_CACHE["qty_map"]

    qty_map = {}
    try:
        rows = load_excel_rows(stock_file, usecols=[0, 1], min_row=2)
        for row in rows:
            item_name = str(row[0]).strip() if row and row[0] not in (None, "") else ""
            qty_val = row[1] if len(row) > 1 else None
            if not item_name or qty_val in (None, ""):
                continue
            try:
                qty = int(float(qty_val))
            except (ValueError, TypeError):
                continue
            if qty <= 0:
                continue
            qty_map[_normalize_text(item_name)] = qty
    except Exception:
        qty_map = {}

    with DATA_CACHE_LOCK:
        STOCK_QTY_CACHE["fingerprint"] = fingerprint
        STOCK_QTY_CACHE["qty_map"] = qty_map
    return qty_map


def _load_training_hierarchy_items():
    qty_map = _load_training_stock_qty_map()
    items = []

    if os.path.exists(MAIN_HIERARCHY_CACHE_JSON):
        try:
            payload = _load_main_hierarchy_payload()
            for entry in payload:
                parent_name = str(entry.get("parent") or "").strip()
                children = entry.get("children", [])
                if not parent_name or not isinstance(children, list):
                    continue
                for child in children:
                    item_name = str(child.get("item_name") or "").strip()
                    if not item_name:
                        continue
                    items.append({
                        "car_model": parent_name,
                        "stock_item_name": item_name,
                        "qty": qty_map.get(_normalize_text(item_name), 0),
                    })
            if items:
                return items
        except Exception:
            pass

    main_file = get_main_file_path()
    if not os.path.exists(main_file):
        return []

    parent_names = {
        _normalize_text(name)
        for name in _load_car_groups_from_cache_or_excel()
        if str(name).strip()
    }
    ignored_labels = {"PARTICULARS", "STOCK SUMMARY", "CLOSING BALANCE", "QUANTITY", ""}
    current_parent = None

    try:
        rows = load_excel_rows(main_file, usecols=[0, 1], min_row=1)
    except Exception:
        return []

    for row in rows:
        item_name = str(row[0]).strip() if row and row[0] not in (None, "") else ""
        if not item_name:
            continue

        item_upper = _normalize_text(item_name)
        if item_upper in ignored_labels:
            continue

        if item_upper in parent_names:
            current_parent = item_name
            continue

        if current_parent is not None:
            items.append({
                "car_model": current_parent,
                "stock_item_name": item_name,
                "qty": qty_map.get(item_upper, 0),
            })

    return items


def get_stock_items_for_training_from_hierarchy(car_full_name):
    if not os.path.exists(MAIN_HIERARCHY_CACHE_JSON):
        return None

    try:
        payload = _load_main_hierarchy_payload()
    except Exception:
        return None

    car_upper = _normalize_text(car_full_name)
    qty_map = _load_training_stock_qty_map()
    for entry in payload:
        parent_name = str(entry.get("parent") or "").strip()
        if _normalize_text(parent_name) != car_upper:
            continue

        children = entry.get("children", [])
        if not isinstance(children, list):
            children = []

        return [
            {
                "car_model": parent_name,
                "stock_item_name": item_name,
                "qty": qty_map.get(_normalize_text(item_name), 0),
            }
            for child in children
            for item_name in [str(child.get("item_name") or "").strip()]
            if item_name
        ]

    return None


def get_stock_items_for_training(car_full_name):
    base_name = extract_car_base_name(car_full_name)
    base_norm = _normalize_text(base_name)
    full_norm = _normalize_text(car_full_name)
    exact_car_norm = _normalize_text(car_full_name)

    items = _load_training_hierarchy_items()
    if not base_norm:
        return items

    exact_items = [
        item for item in items
        if _normalize_text(item.get("car_model")) == exact_car_norm
    ]
    if exact_items:
        return sorted(exact_items, key=lambda item: _normalize_text(item.get("stock_item_name")))

    year_tokens = set(re.findall(r"\b(19\d{2}|20\d{2})\b", full_norm))
    short_year_tokens = {token[-2:] for token in year_tokens}
    variant_tokens = set(re.findall(r"\(([^)]+)\)", full_norm))

    variant_identifiers = set()
    for token in variant_tokens:
        for piece in re.split(r"[^A-Z0-9]+", token):
            piece = piece.strip()
            if not piece or len(piece) < 2:
                continue
            variant_identifiers.add(piece)

    specific_identifiers = set(year_tokens) | short_year_tokens | variant_identifiers

    scored = []
    seen = set()
    for item in items:
        car_model_norm = _normalize_text(item.get("car_model"))
        stock_item_norm = _normalize_text(item.get("stock_item_name"))

        if base_norm not in car_model_norm and base_norm not in stock_item_norm:
            continue

        key = (car_model_norm, stock_item_norm)
        if key in seen:
            continue
        seen.add(key)

        if specific_identifiers:
            matched_identifiers = [token for token in specific_identifiers if token in stock_item_norm]
            if not matched_identifiers:
                continue
            score = len(matched_identifiers)
        else:
            score = 1

        scored.append((score, item))

    if scored:
        scored.sort(key=lambda pair: (-pair[0], _normalize_text(pair[1].get("stock_item_name"))))
        return [pair[1] for pair in scored]

    fallback = []
    fallback_token = next((token for token in base_norm.split() if len(token) > 2), "")
    if not fallback_token:
        return fallback

    for item in items:
        car_model_norm = _normalize_text(item.get("car_model"))
        stock_item_norm = _normalize_text(item.get("stock_item_name"))
        if fallback_token in car_model_norm or fallback_token in stock_item_norm:
            fallback.append(item)

    return fallback


def _check_login_rate_limit(username):
    now = time.time()
    key = str(username or "").strip().lower()
    window_seconds = 300
    max_attempts = 10
    if len(LOGIN_ATTEMPTS) > 500:
        for existing_key in list(LOGIN_ATTEMPTS.keys()):
            existing_attempts = [ts for ts in LOGIN_ATTEMPTS.get(existing_key, []) if now - ts < window_seconds]
            LOGIN_ATTEMPTS[existing_key] = existing_attempts
            if not existing_attempts:
                del LOGIN_ATTEMPTS[existing_key]

    attempts = LOGIN_ATTEMPTS.get(key, [])
    attempts = [ts for ts in attempts if now - ts < window_seconds]
    LOGIN_ATTEMPTS[key] = attempts
    if not attempts:
        del LOGIN_ATTEMPTS[key]
    if len(attempts) >= max_attempts:
        return False
    attempts.append(now)
    LOGIN_ATTEMPTS[key] = attempts
    return True


def _post_tally_with_retry(xml_req):
    response = fetch_from_tally_with_retry(
        TALLY_HTTP_SESSION,
        TALLY_URL,
        xml_req,
        timeout=TALLY_TIMEOUT,
        max_retries=TALLY_RETRY_ATTEMPTS,
        logger=logger,
    )
    return response.text


def _classify_tally_exception(exc):
    message = str(exc or "").lower()
    if isinstance(exc, requests.Timeout) or "timed out" in message:
        return "TALLY_TIMEOUT"
    if isinstance(exc, requests.ConnectionError) or "connection refused" in message or "failed to establish a new connection" in message or "max retries exceeded" in message:
        return "TALLY_UNREACHABLE"
    return "TALLY_ERROR"


def _tally_connection_error_note(exc):
    """Short note to append to a Tally request failure if it looks like a
    connection/timeout error AND multiple Tally windows are currently open;
    "" otherwise.

    Investigated on a real machine (two live tally.exe processes, one
    genuinely listening on port 9000 per netstat, one an inert duplicate
    window): 10/10 real requests succeeded identically with both open, so a
    second Tally.exe isn't actually interfering with anything -- it's only
    worth mentioning as a *possible* explanation once a request has already
    failed for a connection-shaped reason, not treated as a guaranteed cause
    to pre-emptively block on. Non-connection errors (bad XML, no rows
    parsed, etc.) get nothing appended -- a second Tally window had nothing
    to do with those.
    """
    if _classify_tally_exception(exc) not in ("TALLY_TIMEOUT", "TALLY_UNREACHABLE"):
        return ""
    instance_count = _check_multiple_tally_instances()
    if instance_count <= 1:
        return ""
    # Deliberately contains "multiple tally" -- templates/index.html's
    # classifyRefreshIssueBucket() keys its own (unchanged) stale-data
    # banner/hint off that exact substring in status.message/raw_error.
    return f" (Note: multiple Tally windows are open — {instance_count} running; closing the extra one may help.)"


def _raise_with_tally_instance_note(exc):
    """Re-raise exc, appending _tally_connection_error_note()'s context to
    its message when applicable, while preserving its original type so
    downstream isinstance-based classification (_classify_tally_exception)
    still works exactly as before."""
    note = _tally_connection_error_note(exc)
    if not note:
        raise exc
    try:
        raise type(exc)(f"{exc}{note}") from exc
    except TypeError:
        raise Exception(f"{exc}{note}") from exc


def get_main_file_path():
    for path in MAIN_FILE_CANDIDATES:
        if os.path.exists(path):
            return path
    return MAIN_FILE_CANDIDATES[0]


def get_latest_stock_file_path():
    candidates = [
        ITEM_STOCK_FILE_AUTO,
        ITEM_STOCK_FILE_XLSX,
        ITEM_STOCK_FILE,
    ]
    candidates.extend(glob.glob(os.path.join(BASE_DIR, "data", "item stock list.auto.*.xlsx")))
    existing = [
        path for path in candidates
        if os.path.exists(path)
        and not str(path).lower().endswith(".tmp.xlsx")
    ]
    if not existing:
        return None
    return max(existing, key=lambda path: os.path.getmtime(path))


def _scan_images_on_startup():
    try:
        if not app.config["INITIAL_IMAGE_SCAN"]:
            logger.info("Initial image scan disabled by INITIAL_IMAGE_SCAN")
            return
        if not os.path.exists(IMAGE_SCAN_ROOT):
            return
        result = image_scanner.scan_ss_image_folder(IMAGE_SCAN_ROOT)
        print(f"scanned image folder: {result['total_images']} images from {result['total_folders']} folders")
    except Exception as exc:
        print("warning: initial image scan skipped:", exc)


STARTUP_TASKS_STARTED = False

def start_background_startup_tasks():
    global STARTUP_TASKS_STARTED
    if STARTUP_TASKS_STARTED:
        return
    STARTUP_TASKS_STARTED = True

    def _startup_routine():
        try:
            main_file = get_main_file_path()
            if os.path.exists(main_file):
                print("Starting background data load...")
                load_data(refresh_first=False)
            else:
                logger.warning("Skipping background data load because main hierarchy file is missing")
        except Exception:
            logger.exception("Background data load failed")

        try:
            _scan_images_on_startup()
        except Exception:
            logger.exception("Background image scan failed")

        if item_export_enabled:
            schedule_item_export(initial_delay=10)

    thread = threading.Thread(target=_startup_routine, daemon=True)
    thread.start()

    def _startup_full_refresh():
        time.sleep(45)  # give Tally and the app time to be ready
        if FULL_REFRESH_LOCK.acquire(blocking=False):
            try:
                logger.info("Running automatic full refresh on startup")
                run_full_refresh_job()
            finally:
                FULL_REFRESH_LOCK.release()
        else:
            logger.info("Skipping startup full refresh — already running")

    threading.Thread(target=_startup_full_refresh, daemon=True).start()


def _refresh_stock_data():
    global last_refresh_status
    if FULL_REFRESH_LOCK.locked():
        now = datetime.now()
        return {
            "ok": False,
            "busy": True,
            "tally_online": None,
            "status": "full_refresh_in_progress",
            "message": "Full refresh is in progress. Please wait.",
            "timestamp": now.isoformat(),
            "formatted": now.strftime("%d/%m/%Y %H:%M:%S"),
        }
    try:
        export_result = fetch_item_stock_flat()
        load_data()
        msg = "Stock updated successfully"
        if export_result and export_result.get("warning"):
            msg = f"{msg} ({export_result.get('warning')})"
        now = datetime.now()
        last_refresh_status = {
            "success": True,
            "message": msg,
            "timestamp": now.isoformat(),
        }
        timestamp_iso = now.isoformat()
        return {
            "ok": True,
            "tally_online": True,
            "status": "stock updated",
            "message": msg,
            "file": export_result.get("file") if export_result else None,
            "warning": export_result.get("warning") if export_result else None,
            "timestamp": timestamp_iso,
            "formatted": now.strftime("%d/%m/%Y %H:%M:%S"),
        }
    except Exception as exc:
        raw_error = str(exc)
        error_code = _classify_tally_exception(exc)
        if raw_error == MULTIPLE_TALLY_MESSAGE:
            fallback_message = raw_error
        elif error_code == "TALLY_UNREACHABLE":
            fallback_message = f"Tally unreachable at {TALLY_URL}; using the last saved upload. ({raw_error})"
        elif error_code == "TALLY_TIMEOUT":
            fallback_message = f"Tally request timed out; using the last saved upload. ({raw_error})"
        else:
            fallback_message = f"Tally error; using the last saved upload. ({raw_error})"

        now = datetime.now()
        last_refresh_status = {
            "success": False,
            "message": fallback_message,
            "timestamp": now.isoformat(),
            "error_code": error_code,
            "raw_error": raw_error,
        }
        timestamp_iso = now.isoformat()
        return {
            "ok": True,
            "tally_online": False,
            "status": "using_last_saved_upload",
            "message": fallback_message,
            "error_code": error_code,
            "raw_error": raw_error,
            "file": get_latest_stock_file_path(),
            "warning": fallback_message,
            "timestamp": timestamp_iso,
            "formatted": now.strftime("%d/%m/%Y %H:%M:%S"),
                }

# ---------- Tally export logic ----------

def fetch_item_stock_flat():
    """Export item stock from Tally and save as clean Excel.

    Uses a single TDL collection walk over StockItem masters (NAME +
    CLOSINGBALANCE). This replaced a three-request approach whose
    detailed+exploded Stock Summary made Tally render its entire stock
    tree on every cycle (~15-20s of engine work on production data,
    measured via the System panel's Tally Performance Test), visibly
    stalling Tally Prime. A collection returns only items -- never
    group rows -- so the old group-name filtering requests are
    unnecessary. Output format is unchanged.
    """
    def _norm(text: str) -> str:
        return re.sub(r"\s+", " ", str(text or "").strip()).upper()

    def _fetch_item_rows():
        xml_req = '''<ENVELOPE>
    <HEADER>
        <VERSION>1</VERSION>
        <TALLYREQUEST>Export Data</TALLYREQUEST>
        <TYPE>Collection</TYPE>
        <ID>Item Names With Closing</ID>
    </HEADER>
    <BODY>
        <DESC>
            <STATICVARIABLES>
                <SVEXPORTFORMAT>$$SysName:XML</SVEXPORTFORMAT>
            </STATICVARIABLES>
            <TDL>
                <TDLMESSAGE>
                    <COLLECTION NAME="Item Names With Closing" ISMODIFY="No">
                        <TYPE>StockItem</TYPE>
                        <FETCH>NAME, CLOSINGBALANCE</FETCH>
                    </COLLECTION>
                </TDLMESSAGE>
            </TDL>
        </DESC>
    </BODY>
</ENVELOPE>'''.strip()

        text = _post_tally_with_retry(xml_req)
        if "<LINEERROR>" in text or "<RESPONSE>Error" in text or "Unknown Request" in text:
            raise Exception(f"Tally returned error for item stock collection: {text[:300]}")

        text = _sanitize_tally_xml(text)
        root = ET.fromstring(text)
        parsed_rows = []
        for item_node in root.findall('.//COLLECTION/STOCKITEM'):
            raw_name = item_node.attrib.get('NAME', '').strip()
            if not raw_name:
                name_node = item_node.find('.//NAME')
                raw_name = name_node.text.strip() if (name_node is not None and name_node.text) else ''
            bal_node = item_node.find('.//CLOSINGBALANCE')
            qty_text = bal_node.text if (bal_node is not None and bal_node.text) else "0"
            match = re.search(r"-?\d+", qty_text)
            qty = int(match.group()) if match else 0
            if raw_name and qty > 0:
                parsed_rows.append({"item_name": raw_name, "qty": qty})
        return parsed_rows

    def _load_main_name_set():
        main_file = get_main_file_path()
        if not os.path.exists(main_file):
            return set()

        try:
            values = load_excel_column(main_file, col_index=0, min_row=1)
        except Exception:
            return set()

        names = set()
        for raw in values:
            normalized = _norm(raw)
            if normalized and normalized not in {"PARTICULARS", "STOCK SUMMARY", "CLOSING BALANCE", "QUANTITY"}:
                names.add(normalized)
        return names

    print("requesting flat item stock list from Tally...")
    try:
        export_started = time.perf_counter()
        rows = _fetch_item_rows()
        main_name_set = _load_main_name_set()

        # Optional strict alignment with main hierarchy names, when available.
        if main_name_set:
            aligned_rows = [r for r in rows if _norm(r["item_name"]) in main_name_set]
            if aligned_rows:
                rows = aligned_rows

        if not rows:
            raise Exception("No stock rows parsed from Tally item collection")

        seen = {}
        for r in rows:
            seen[r["item_name"]] = r
        deduped = list(seen.values())

        try:
            with open(ITEM_STOCK_CACHE_JSON, "w", encoding="utf-8") as handle:
                json.dump(deduped, handle, ensure_ascii=False)
        except Exception:
            pass

        threading.Thread(target=_write_stock_excel, args=(deduped,), daemon=True).start()
        logger.info(
            "Item stock export completed in %.1fs across 1 Tally request (%d items)",
            time.perf_counter() - export_started, len(deduped),
        )
        return {"rows": len(deduped), "file": ITEM_STOCK_FILE_AUTO, "warning": None}
    except Exception as exc:
        print(f"item stock export failed: {str(exc)}")
        _raise_with_tally_instance_note(exc)


def fetch_car_master_from_tally():
    def _norm(text: str) -> str:
        return re.sub(r"\s+", " ", str(text or "").strip()).upper()

    xml_req = '''<ENVELOPE>
  <HEADER>
    <VERSION>1</VERSION>
    <TALLYREQUEST>Export Data</TALLYREQUEST>
    <TYPE>Collection</TYPE>
    <ID>List of Stock Groups</ID>
  </HEADER>
  <BODY>
    <DESC>
      <STATICVARIABLES>
        <SVEXPORTFORMAT>$$SysName:XML</SVEXPORTFORMAT>
      </STATICVARIABLES>
    </DESC>
  </BODY>
</ENVELOPE>'''.strip()

    try:
        text = _post_tally_with_retry(xml_req)
    except Exception as exc:
        _raise_with_tally_instance_note(exc)

    if "<LINEERROR>" in text or "Unknown Request" in text:
        raise Exception(f"Tally returned error for stock group collection: {text[:300]}")

    text = _sanitize_tally_xml(text)
    root = ET.fromstring(text)
    car_names = set()
    for item_node in root.findall('.//COLLECTION/STOCKGROUP'):
        raw_name = item_node.attrib.get('NAME', '')
        if not raw_name:
            name_node = item_node.find('.//NAME')
            raw_name = name_node.text if (name_node is not None and name_node.text) else ''
        norm_name = _norm(raw_name)
        if norm_name:
            car_names.add(norm_name)

    return sorted(car_names)


def _car_names_with_real_children():
    """Return the set of normalized car/parent names that have at least one
    real child entry in main_hierarchy.json, or None if that file is
    missing/unreadable/empty (in which case callers should skip filtering
    rather than risk emptying the whole dropdown).

    A car can appear in car_master.json's raw Tally Stock Group list (e.g.
    fetched before a car was deleted from Tally, or simply outside the
    hierarchy fetch for some other reason) while having no corresponding
    entry here at all -- that car has zero real designs and must not show
    in the dropdown. This is distinct from a car that DOES have real
    children here but where every child currently has zero stock (e.g.
    "ALCAZAR 8 ARMS 2024**** N1") -- that case must still show, so the
    check below only requires a non-empty children list, not any qty > 0.
    """
    if not os.path.exists(MAIN_HIERARCHY_CACHE_JSON):
        return None
    try:
        with open(MAIN_HIERARCHY_CACHE_JSON, "r", encoding="utf-8") as handle:
            payload = json.load(handle)
    except Exception:
        return None
    if not isinstance(payload, list):
        return None

    names = set()
    for entry in payload:
        parent_name = str(entry.get("parent") or "").strip()
        children = entry.get("children")
        if parent_name and isinstance(children, list) and children:
            names.add(normalize_text(parent_name))
    return names or None


def _load_car_groups_from_cache_or_excel():
    car_groups = []

    if os.path.exists(CAR_MASTER_CACHE_JSON):
        try:
            with open(CAR_MASTER_CACHE_JSON, "r", encoding="utf-8") as handle:
                payload = json.load(handle)
            if isinstance(payload, list):
                car_groups = [str(value).strip() for value in payload if str(value).strip()]
        except Exception:
            car_groups = []

    if not car_groups:
        try:
            values = load_excel_column(CAR_FILE, col_index=0, min_row=1)
            car_groups = [str(value).strip() for value in values if str(value).strip()]
        except Exception as exc:
            raise Exception(f"Failed to load car master list: {exc}")

    # ============================================================
    # !! CRITICAL - CAR MODEL RANGE FILTER - READ THIS FIRST !!
    # ============================================================
    # The list of car models loaded from Tally is trimmed to only
    # the range between start_marker and end_marker below.
    #
    # start_marker = "ACCESSORIES"
    # end_marker   = "ZS - EV FOOT MAT"
    #
    # IF THE CAR DROPDOWN IS MISSING MODELS OR SHOWING WRONG CARS:
    #   1. Open Tally Prime
    #   2. Check the Stock Groups list
    #   3. Confirm these two group names still exist exactly as written
    #   4. If either name changed, update start_marker and end_marker here
    #   5. Then run a Full Refresh from the admin UI
    #
    # IF THIS FILTER IS REMOVED, ALL TALLY STOCK GROUPS WILL APPEAR
    # IN THE DROPDOWN INCLUDING INTERNAL/ACCOUNTING GROUPS.
    # ============================================================
    start_marker = "ACCESSORIES"
    end_marker = "ZS - EV FOOT MAT"
    if start_marker in car_groups and end_marker in car_groups:
        start_idx = car_groups.index(start_marker)
        end_idx = car_groups.index(end_marker)
        if start_idx <= end_idx:
            car_groups = car_groups[start_idx : end_idx + 1]
        else:
            car_groups = car_groups[end_idx : start_idx + 1]

    return car_groups


def save_car_master_to_file(car_names):
    temp_file = CAR_FILE + ".tmp.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    for car_name in car_names:
        ws.append([car_name])
    wb.save(temp_file)
    os.replace(temp_file, CAR_FILE)
    with open(CAR_MASTER_CACHE_JSON, "w", encoding="utf-8") as handle:
        json.dump(car_names, handle, ensure_ascii=False)
    print(f"wrote {len(car_names)} car names to {CAR_FILE}")


def fetch_main_hierarchy_from_tally():
    def _norm(text: str) -> str:
        return re.sub(r"\s+", " ", str(text or "").strip()).upper()

    def _node_text(node, tag_name: str) -> str:
        child = node.find(f".//{tag_name}")
        return child.text.strip() if (child is not None and child.text) else ""

    def _load_existing_flat_rows():
        main_file = get_main_file_path()
        if not os.path.exists(main_file):
            return None

        try:
            rows_data = load_excel_rows(main_file, usecols=[0, 1], min_row=1)
        except Exception:
            logger.exception("Failed to read existing main hierarchy for fallback")
            return None

        flat_rows = []
        for row in rows_data:
            item_name = str(row[0]).strip() if row and row[0] not in (None, "") else ""
            qty = row[1] if len(row) > 1 else ""
            qty_text = str(qty or "").strip().upper()
            if item_name.upper() == "PARTICULARS" and qty_text == "QUANTITY":
                continue
            flat_rows.append({"item_name": item_name, "qty": "" if qty is None else qty})
        return flat_rows or None

    def _fallback_existing(reason: str):
        existing_rows = _load_existing_flat_rows()
        if existing_rows:
            logger.warning("%s Keeping existing main hierarchy unchanged.", reason)
            return existing_rows
        raise Exception(reason)

    def _fetch_stock_items_with_parent():
        xml_req = '''<ENVELOPE>
    <HEADER>
        <VERSION>1</VERSION>
        <TALLYREQUEST>Export Data</TALLYREQUEST>
        <TYPE>Collection</TYPE>
        <ID>List of Stock Items with Parent</ID>
    </HEADER>
    <BODY>
        <DESC>
            <STATICVARIABLES>
                <SVEXPORTFORMAT>$$SysName:XML</SVEXPORTFORMAT>
            </STATICVARIABLES>
            <TDL>
                <TDLMESSAGE>
                    <COLLECTION NAME="List of Stock Items with Parent" ISMODIFY="No">
                        <TYPE>StockItem</TYPE>
                        <FETCH>NAME, PARENT</FETCH>
                    </COLLECTION>
                </TDLMESSAGE>
            </TDL>
        </DESC>
    </BODY>
</ENVELOPE>'''.strip()

        text = _post_tally_with_retry(xml_req)
        if "<LINEERROR>" in text or "<RESPONSE>Error" in text or "Unknown Request" in text:
            raise Exception(f"TDL collection request failed for stock items with parent: {text[:300]}")

        text = _sanitize_tally_xml(text)
        root = ET.fromstring(text)
        items = []
        for item_node in root.findall(".//COLLECTION/STOCKITEM"):
            item_name = item_node.attrib.get("NAME", "").strip() or _node_text(item_node, "NAME")
            parent_name = item_node.attrib.get("PARENT", "").strip() or _node_text(item_node, "PARENT")
            if item_name and parent_name:
                items.append((item_name, parent_name))
        return items

    def _fetch_current_qty_map():
        xml_req = '''<ENVELOPE>
    <HEADER>
        <VERSION>1</VERSION>
        <TALLYREQUEST>Export</TALLYREQUEST>
        <TYPE>Data</TYPE>
        <ID>Stock Summary</ID>
    </HEADER>
    <BODY>
        <DESC>
            <STATICVARIABLES>
                <SVEXPORTFORMAT>$$SysName:XML</SVEXPORTFORMAT>
                <SVSTOCKGROUP>Primary</SVSTOCKGROUP>
                <ISDETAILED>Yes</ISDETAILED>
                <EXPLODEFLAG>Yes</EXPLODEFLAG>
                <SVSHOWALLITEMS>Yes</SVSHOWALLITEMS>
                <SVSHOWZEROBALANCES>Yes</SVSHOWZEROBALANCES>
            </STATICVARIABLES>
        </DESC>
    </BODY>
</ENVELOPE>'''.strip()

        text = _post_tally_with_retry(xml_req)
        if "<LINEERROR>" in text or "<RESPONSE>Error" in text or "Unknown Request" in text:
            raise Exception(f"Tally returned error while fetching detailed stock quantities: {text[:300]}")

        text = _sanitize_tally_xml(text)
        root = ET.fromstring(text)
        names = root.findall(".//DSPACCNAME")
        stocks = root.findall(".//DSPSTKINFO")

        qty_map = {}
        for name_node, stock_node in zip(names, stocks):
            display_node = name_node.find("DSPDISPNAME")
            qty_node = stock_node.find(".//DSPCLQTY")
            item_name = display_node.text.strip() if (display_node is not None and display_node.text) else ""
            qty_text = qty_node.text if (qty_node is not None and qty_node.text) else "0"
            match = re.search(r"-?\d+", qty_text)
            qty = int(match.group()) if match else 0
            if item_name and qty > 0:
                qty_map[_norm(item_name)] = qty
        return qty_map

    stock_items = []
    try:
        stock_items = _fetch_stock_items_with_parent()
    except ET.ParseError:
        raise
    except Exception as exc:
        return _fallback_existing(f"{exc}{_tally_connection_error_note(exc)}")

    if not stock_items:
        return _fallback_existing("Stock item master collection returned no items for main hierarchy refresh.")

    try:
        qty_map = _fetch_current_qty_map()
    except Exception as exc:
        _raise_with_tally_instance_note(exc)
    allowed_parents = _load_car_groups_from_cache_or_excel()
    allowed_parent_lookup = {_norm(parent_name): parent_name for parent_name in allowed_parents}

    grouped_items = defaultdict(list)
    for item_name, parent_name in stock_items:
        parent_upper = _norm(parent_name)
        canonical_parent = allowed_parent_lookup.get(parent_upper)
        if canonical_parent:
            grouped_items[canonical_parent].append(item_name)

    if not grouped_items:
        return _fallback_existing("Stock item master collection returned no items in the configured car range.")

    flat_rows = []
    for parent_name in allowed_parents:
        child_names = grouped_items.get(parent_name, [])
        if not child_names:
            continue

        child_rows = []
        parent_qty = 0
        for item_name in child_names:
            qty = qty_map.get(_norm(item_name), 0)
            parent_qty += qty
            child_rows.append({"item_name": item_name, "qty": qty})

        flat_rows.append({"item_name": parent_name, "qty": parent_qty})
        flat_rows.extend(child_rows)
        flat_rows.append({"item_name": "", "qty": ""})

    if not flat_rows:
        return _fallback_existing("No parent-child hierarchy rows were built from the stock item master collection.")

    return flat_rows


def _build_main_hierarchy_structure(flat_rows):
    parent_names = {
        normalize_text(name)
        for name in _load_car_groups_from_cache_or_excel()
        if str(name).strip()
    }
    structured_rows = []
    current_parent = None

    for row in flat_rows or []:
        item_name = str(row.get("item_name") or "").strip()
        qty = row.get("qty") or 0
        try:
            qty = int(qty)
        except (TypeError, ValueError):
            qty = 0

        if not item_name:
            continue

        item_name_upper = normalize_text(item_name)
        if item_name_upper in parent_names:
            current_parent = {
                "parent": item_name,
                "parent_qty": qty,
                "children": [],
            }
            structured_rows.append(current_parent)
            continue

        if current_parent is not None:
            current_parent["children"].append({
                "item_name": item_name,
                "qty": qty,
            })

    return structured_rows


def save_main_hierarchy_to_file(flat_rows):
    main_file = get_main_file_path()
    temp_file = main_file + ".tmp.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["PARTICULARS", "QUANTITY"])
    for row in flat_rows:
        ws.append([row.get("item_name", ""), row.get("qty", 0)])
    wb.save(temp_file)
    os.replace(temp_file, main_file)

    structured_rows = _build_main_hierarchy_structure(flat_rows)
    with open(MAIN_HIERARCHY_CACHE_JSON, "w", encoding="utf-8") as handle:
        json.dump(structured_rows, handle, ensure_ascii=False)

    print(f"wrote {len(flat_rows)} hierarchy rows to {main_file}")


def _write_stock_excel(deduped):
    EXPORT_LOCK.acquire()
    try:
        temp_file = ITEM_STOCK_FILE_AUTO + ".tmp.xlsx"
        out_file = ITEM_STOCK_FILE_AUTO
        warning = None

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["item_name", "qty"])
        for r in deduped:
            ws.append([r["item_name"], r["qty"]])
        wb.save(temp_file)

        replaced = False
        for attempt in range(MAX_RETRY_ATTEMPTS):
            try:
                os.replace(temp_file, ITEM_STOCK_FILE_AUTO)
                replaced = True
                out_file = ITEM_STOCK_FILE_AUTO
                break
            except PermissionError:
                time.sleep(0.1 * (2 ** attempt))
            except OSError as exc:
                if getattr(exc, "winerror", None) == 5:
                    time.sleep(0.1 * (2 ** attempt))
                else:
                    raise

        if not replaced:
            warning = "auto file locked; wrote to alternate file"
            try:
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.append(["item_name", "qty"])
                for r in deduped:
                    ws.append([r["item_name"], r["qty"]])
                wb.save(ITEM_STOCK_FILE_XLSX)
                out_file = ITEM_STOCK_FILE_XLSX
            except Exception:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                out_file = os.path.join(BASE_DIR, "data", f"item stock list.auto.{timestamp}.xlsx")
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.append(["item_name", "qty"])
                for r in deduped:
                    ws.append([r["item_name"], r["qty"]])
                wb.save(out_file)

        try:
            if os.path.exists(temp_file):
                os.remove(temp_file)
        except Exception:
            pass

        print(f"exported item stock list to {out_file} ({len(deduped)} rows)")
        if warning:
            print(warning)
        return {"rows": len(deduped), "file": out_file, "warning": warning}
    except Exception:
        logger.exception("background item stock excel write failed")
    finally:
        EXPORT_LOCK.release()


def schedule_item_export(initial_delay: int = 0):
    global item_export_timer, last_refresh_status

    def _export_job():
        global item_export_timer

        if FULL_REFRESH_LOCK.locked():
            logger.info("Skipping scheduled export because full refresh is in progress")
            item_export_timer = threading.Timer(ITEM_EXPORT_INTERVAL, schedule_item_export)
            item_export_timer.daemon = True
            item_export_timer.start()
            return

        if not EXPORT_LOCK.acquire(blocking=False):
            logger.warning("Skipping scheduled export because previous export is still running")
            item_export_timer = threading.Timer(ITEM_EXPORT_INTERVAL, schedule_item_export)
            item_export_timer.daemon = True
            item_export_timer.start()
            return

        try:
            export_result = fetch_item_stock_flat()
            try:
                load_data()
            except Exception:
                pass
            msg = "Stock updated successfully"
            if export_result and export_result.get("warning"):
                msg = f"{msg} ({export_result.get('warning')})"
            last_refresh_status = {
                "success": True,
                "message": msg,
                "timestamp": datetime.now().isoformat()
            }
        except Exception as exc:
            logger.exception("scheduled item stock export failed")
            last_refresh_status = {
                "success": False,
                "message": str(exc),
                "timestamp": datetime.now().isoformat()
            }
        finally:
            EXPORT_LOCK.release()

        item_export_timer = threading.Timer(ITEM_EXPORT_INTERVAL, schedule_item_export)
        item_export_timer.daemon = True
        item_export_timer.start()

    if initial_delay and initial_delay > 0:
        item_export_timer = threading.Timer(initial_delay, _export_job)
        item_export_timer.daemon = True
        item_export_timer.start()
        return

    _export_job()


# refresh status (auto + manual item stock export)
last_refresh_status = {"success": False, "message": "Not yet run", "timestamp": None}

# the actual start of the timer happens later, once all helper functions
# (especially load_data) are defined.  see bottom of this file.

# ----------------------------
# Hierarchical parser for Tally export
# ----------------------------
def parse_flat_tally(file_path):
    """Load all designs from Tally export as a flat list (no hierarchy parsing).
    
    Simply reads all valid rows and returns them as a list for matching to cars.
    
    Returns:
        list: [{"design": str, "raw": str, "qty": int}, ...]
    """
    designs = []
    ignored_labels = {"PARTICULARS", "STOCK SUMMARY", "CLOSING BALANCE", "QUANTITY", ""}

    print(f"Reading Tally rows from {file_path}")
    rows = load_excel_rows(file_path, usecols=[0, 1], min_row=2)
    for raw0, col1_val in rows:
        col0 = str(raw0).strip() if raw0 not in (None, "") else ""
        if not col0 or _normalize_text(col0) in ignored_labels:
            continue

        qty = None
        if col1_val not in (None, ""):
            try:
                qty = int(float(col1_val))
            except (ValueError, TypeError):
                pass

        if qty is not None and qty > 0:
            designs.append({
                "raw": col0,
                "design": col0,
                "qty": qty,
            })

    print(f"Loaded {len(designs)} valid designs")
    return designs



# In-memory cache (module-level globals)
CAR_GROUPS = []
CAR_DESIGN_MAP = {}
TALLY_HIERARCHY = {}  # Store raw Tally hierarchy: {parent_name: [child_items]}


def load_data(refresh_first: bool = False):
    """Populate in-memory caches for dual-file workflow.

    - car master list.xls provides dropdown car models
    - main.xlsx/main.xls provides parent-child hierarchy only
    - item stock list.xls provides live quantities
    """
    global CAR_GROUPS, CAR_DESIGN_MAP, TALLY_HIERARCHY, PARENT_NAME_SET

    CAR_DESIGN_MAP = {}
    CAR_GROUPS = []
    TALLY_HIERARCHY = {}
    PARENT_NAME_SET = set()
    _invalidate_runtime_caches()

    # optionally pull a fresh export from Tally
    if refresh_first:
        try:
            fetch_item_stock_flat()
        except Exception as exc:
            print("warning: could not refresh stock from Tally:", exc)

    # ---- Load car groups from car master list ----
    try:
        raw_car_groups = _load_car_groups_from_cache_or_excel()
        # PARENT_NAME_SET stays built from the full, unfiltered list -- it's
        # also used as a row-boundary marker when scanning main.xlsx (see
        # _find_children_by_qty), which is unrelated to what shows in the
        # dropdown and must not change here.
        PARENT_NAME_SET = {normalize_text(name) for name in raw_car_groups if str(name).strip()}

        real_children_names = _car_names_with_real_children()
        if real_children_names is not None:
            CAR_GROUPS = [
                car for car in raw_car_groups
                if normalize_text(car) in real_children_names
            ]
        else:
            CAR_GROUPS = raw_car_groups

        print(f"[OK] Loaded {len(CAR_GROUPS)} car models from dropdown")
    except Exception as exc:
        print(f"[ERROR] Error loading car master list: {exc}")
        return

    # ---- Load designs from Tally export as flat list ----
    try:
        main_file = get_main_file_path()
        if not os.path.exists(main_file):
            print(f"[ERROR] Main file not found. Tried: {', '.join(MAIN_FILE_CANDIDATES)}")
            print(f"  Designs will be empty until you export from Tally")
            return
        
        # Load all designs from Tally as a flat list
        all_designs = parse_flat_tally(main_file)
        design_index = defaultdict(list)
        for index, design_item in enumerate(all_designs):
            for token in normalize_text(design_item["raw"]).split():
                if len(token) > 2:
                    design_index[token].append(index)

        for car in CAR_GROUPS:
            base = normalize_text(extract_car_base_name(car) or car)
            candidate_indices = set()
            for token in base.split():
                if len(token) > 2:
                    for index in design_index.get(token, []):
                        candidate_indices.add(index)
            CAR_DESIGN_MAP[car] = [all_designs[index] for index in sorted(candidate_indices)]

        total_designs = sum(len(v) for v in CAR_DESIGN_MAP.values())
        print(f"Matched {len(CAR_DESIGN_MAP)} cars to {total_designs} designs")
    
    except Exception as exc:
        print(f"[ERROR] Error loading Tally designs: {exc}")
        import traceback
        traceback.print_exc()


def ensure_data_loaded():
    """Load cached data on demand if startup initialization did not run."""
    if CAR_GROUPS:
        return None

    try:
        load_data(refresh_first=False)
    except Exception as exc:
        return str(exc)

    if not CAR_GROUPS:
        return "No car models loaded from car master list"

    return None



# ----------------------------
# API routes
# ----------------------------
@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "GET":
        if _current_role() is not None:
            return redirect(url_for("home"))
        return render_template(
            "login.html",
            error=None,
            next_url=_safe_next_url(request.args.get("next")),
        )

    username = (request.form.get("username") or "").strip()
    access_code = (request.form.get("access_code") or "").strip()
    next_url = _safe_next_url(request.form.get("next") or request.args.get("next"))

    if not _check_login_rate_limit(username):
        return render_template(
            "login.html",
            error="Too many login attempts. Please try again later.",
            next_url=next_url,
        ), 429

    user_record = db.authenticate_user(username, access_code)
    if not user_record:
        return render_template(
            "login.html",
            error="Invalid username or access code.",
            next_url=next_url,
        ), 401

    if user_record.get("role") == "customer" and not user_record.get("is_active", 1):
        return render_template(
            "login.html",
            error="This account has been paused. Contact the administrator.",
            next_url=next_url,
        ), 403

    db.update_last_login(user_record["id"])

    session.clear()
    session["user_id"] = user_record["id"]
    session["username"] = user_record["username"]
    session["role"] = user_record["role"]
    return redirect(next_url)


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


@app.route("/", methods=["GET"])
def home():
    # serve the frontend page
    # pass a couple of flags so the UI can indicate whether auto-export is running
    return render_template(
        "index.html",
        auto_export_enabled=item_export_enabled,
        refresh_interval=ITEM_EXPORT_INTERVAL,
        last_refresh_status=last_refresh_status,
    )
@app.route("/health")
def health():
    db_ok = True
    db_error = None
    try:
        stats = db.get_mapping_stats()
    except Exception as exc:
        db_ok = False
        stats = None
        db_error = str(exc)
        logger.exception("Health check database failure")

    return jsonify({
        "status": "ok" if db_ok else "degraded",
        "database": {"ok": db_ok, "error": db_error},
        "tally_url": TALLY_URL,
        "auto_export_enabled": item_export_enabled,
        "mapping_stats": stats,
    }), (200 if db_ok else 503)



@app.route("/cars")
def cars():
    # return list of car models (column A of car master Excel)
    load_error = ensure_data_loaded()
    if load_error:
        return jsonify({"error": load_error}), 500
    print("/cars endpoint called; returning", len(CAR_GROUPS), "models")
    return jsonify(CAR_GROUPS)


def _find_children_by_qty(car_name: str):
    """Search MAIN_FILE for the row matching `car_name` and return all
    subsequent rows until the summed quantity equals the parent's quantity.

    This implements the new strategy proposed by the user:
    > let the option selected from the dropdown be the key/parent,
    > search it in the main excel sheet, beside the total quantity is known,
    > print the lines below it while add the quantities of each of those lines,
    > print until the total quantity matches the quantity of the key.
    """

    def _norm(text: str) -> str:
        return re.sub(r"\s+", " ", str(text or "").strip()).upper()

    if not PARENT_NAME_SET:
        logger.warning("Skipping child lookup because parent data has not loaded yet")
        return False, []

    def _to_int(value):
        if value is None:
            return 0
        match = re.search(r"-?\d+", str(value))
        return int(match.group()) if match else 0

    def _load_stock_qty_map_cached():
        json_file = ITEM_STOCK_CACHE_JSON
        if os.path.exists(json_file):
            try:
                with open(json_file, "r", encoding="utf-8") as handle:
                    payload = json.load(handle)
                qty_map = {}
                for row in payload if isinstance(payload, list) else []:
                    item_name = str(row.get("item_name") or "").strip()
                    try:
                        qty = int(row.get("qty") or 0)
                    except (TypeError, ValueError):
                        continue
                    if item_name and qty > 0:
                        qty_map[_norm(item_name)] = qty
                if qty_map:
                    with DATA_CACHE_LOCK:
                        STOCK_QTY_CACHE["fingerprint"] = _file_fingerprint(json_file)
                        STOCK_QTY_CACHE["qty_map"] = qty_map
                    return qty_map
            except Exception:
                pass

        stock_file = get_latest_stock_file_path()
        fingerprint = _file_fingerprint(stock_file)
        if fingerprint is None:
            return {}

        with DATA_CACHE_LOCK:
            if STOCK_QTY_CACHE["fingerprint"] == fingerprint:
                return STOCK_QTY_CACHE["qty_map"]

        qty_map = {}
        try:
            rows = load_excel_rows(stock_file, usecols=[0, 1], min_row=2)
            for row in rows:
                item_name = str(row[0]).strip() if row and row[0] not in (None, "") else ""
                qty_val = row[1] if len(row) > 1 else None
                if not item_name or qty_val in (None, ""):
                    continue
                try:
                    qty = int(float(qty_val))
                except (ValueError, TypeError):
                    continue
                if qty <= 0:
                    continue
                qty_map[_norm(item_name)] = qty
        except Exception:
            qty_map = {}

        with DATA_CACHE_LOCK:
            STOCK_QTY_CACHE["fingerprint"] = fingerprint
            STOCK_QTY_CACHE["qty_map"] = qty_map
        return qty_map

    def _lookup_from_hierarchy_json(car_name: str):
        if not os.path.exists(MAIN_HIERARCHY_CACHE_JSON):
            return False, None

        try:
            with open(MAIN_HIERARCHY_CACHE_JSON, "r", encoding="utf-8") as handle:
                payload = json.load(handle)
        except Exception:
            return False, None

        car_upper = _norm(car_name)
        for entry in payload if isinstance(payload, list) else []:
            parent_name = str(entry.get("parent") or "").strip()
            if _norm(parent_name) != car_upper:
                continue

            children = entry.get("children", [])
            if not isinstance(children, list):
                children = []

            stock_qty_map = _load_stock_qty_map_cached()
            live_children = []
            for child in children:
                item_name = str(child.get("item_name") or "").strip()
                if not item_name:
                    continue
                live_qty = stock_qty_map.get(_norm(item_name), 0)
                if live_qty > 0:
                    live_children.append({"raw": item_name, "design": item_name, "qty": live_qty})
            return True, live_children

        return False, None

    found_in_json, json_children = _lookup_from_hierarchy_json(car_name)
    if found_in_json:
        return True, json_children

    def _load_main_rows_cached():
        main_file = get_main_file_path()
        fingerprint = _file_fingerprint(main_file)
        if fingerprint is None:
            return [], {}

        with DATA_CACHE_LOCK:
            if MAIN_ROWS_CACHE["fingerprint"] == fingerprint:
                return MAIN_ROWS_CACHE["rows"], MAIN_ROWS_CACHE["exact_index"]

        rows = []
        exact_index = {}
        ignore = {"PARTICULARS", "STOCK SUMMARY", "CLOSING BALANCE", "QUANTITY", ""}
        
        try:
            rows_data = load_excel_rows(main_file, usecols=[0, 1], min_row=1)
            for row in rows_data:
                name = str(row[0]).strip() if row and row[0] not in (None, "") else ""
                if not name:
                    continue
                name_upper = _norm(name)
                if name_upper in ignore:
                    continue
                qty = _to_int(row[1] if len(row) > 1 else None)
                rows.append((name, name_upper, qty))
                exact_index.setdefault(name_upper, len(rows) - 1)
        except Exception:
            return [], {}

        with DATA_CACHE_LOCK:
            MAIN_ROWS_CACHE["fingerprint"] = fingerprint
            MAIN_ROWS_CACHE["rows"] = rows
            MAIN_ROWS_CACHE["exact_index"] = exact_index
        return rows, exact_index

    rows, exact_index = _load_main_rows_cached()
    if not rows:
        return False, []

    parent_idx = None
    car_upper = _norm(car_name)
    matched_parent_upper = ""
    parent_qty_total = 0

    exact_match_idx = exact_index.get(car_upper)
    if exact_match_idx is not None:
        parent_idx = exact_match_idx
        _, matched_parent_upper, parent_qty_total = rows[parent_idx]
    else:
        for idx, (_, name_upper, qty_total) in enumerate(rows):
            if car_upper in name_upper:
                parent_idx = idx
                matched_parent_upper = name_upper
                parent_qty_total = qty_total
                break

    if parent_idx is None:
        return False, []

    stock_qty_map = _load_stock_qty_map_cached()

    children = []
    running_qty = 0
    for name, upper_name, _ in rows[parent_idx + 1 :]:

        if parent_qty_total > 0 and running_qty >= parent_qty_total:
            break

        # stop at next parent group to keep only this section's child items
        if upper_name in PARENT_NAME_SET:
            break

        # never return parent labels as designs
        if upper_name == matched_parent_upper:
            continue

        # Lookup quantity from item stock list
        qty = stock_qty_map.get(upper_name, 0)
        if qty > 0:
            children.append({"raw": name, "design": name, "qty": qty})
            running_qty += qty

            if parent_qty_total > 0 and running_qty >= parent_qty_total:
                break
    # Car was found as a real parent row (matched exactly or by substring
    # above), just its own listed children summed to zero currently-in-stock
    # items -- this is the same honest "found, but empty" state as a car
    # with real hierarchy children that all happen to be zero-stock right
    # now (e.g. "ALCAZAR 8 ARMS 2024**** N1"), not a "car doesn't exist"
    # state, so found stays True.
    return True, children


def _coerce_limit(value, default=None, default_value=None):
    # Accept both `default` and legacy `default_value` keyword to remain
    # compatible with older call sites.
    if default_value is not None and default is None:
        default = default_value
    try:
        v = int(value) if value is not None else default
        return max(1, min(v, MAX_IMAGE_RESPONSE_LIMIT)) if v is not None else None
    except (TypeError, ValueError):
        return max(1, min(default, MAX_IMAGE_RESPONSE_LIMIT)) if default is not None else None


def _parse_bool(value):
    text = str(value or "").strip().lower()
    return text in {"1", "true", "yes", "on"}


@app.route("/designs")
def designs():
    car = request.args.get("car")
    if not car:
        return jsonify([])

    load_error = ensure_data_loaded()
    if load_error:
        return jsonify({"error": load_error}), 500

    # quantity-sum scanning of the export. A car with no real hierarchy
    # entry at all (deleted from Tally, never in the range fetch, etc.)
    # and a car found with zero currently-in-stock children both return an
    # honest empty list here -- CAR_DESIGN_MAP's loose token matching is no
    # longer used as a fallback, since it silently matched unrelated cars
    # sharing a generic word (e.g. "MAT") and returned hundreds of wrong
    # designs for a genuinely empty car.
    _found, children = _find_children_by_qty(car)
    if children:
        return jsonify(_build_design_payload(children))

    return jsonify([])


@app.route("/api/get_all_items_for_car")
@admin_required
def get_all_items_for_car():
    """Admin-only, qty-agnostic sibling of /designs -- backs the "Assign
    Category" selection flow, which must be able to tag zero-stock and
    not-yet-image-mapped items too, not just what /designs shows customers.
    Reuses get_stock_items_for_training_from_hierarchy() (the same "every
    child under this car regardless of stock" lookup Training Mode's fixed
    item list is built from) rather than re-deriving that list."""
    car = request.args.get("car")
    if not car:
        return jsonify([])

    load_error = ensure_data_loaded()
    if load_error:
        return jsonify({"error": load_error}), 500

    items = get_stock_items_for_training_from_hierarchy(car)
    if not items:
        return jsonify([])

    designs_input = [
        {"raw": item["stock_item_name"], "design": item["stock_item_name"], "qty": item.get("qty", 0)}
        for item in items
    ]
    return jsonify(_build_design_payload(designs_input))


def _sort_queue_bucket(bucket, remaining_key):
    bucket.sort(key=lambda entry: (-entry[remaining_key], entry["car"]))


def _queue_item_name_fields(stats, missing_items_key, remaining):
    """Builds the hybrid display's per-car {"items": [...], "items_more": n}
    fields (Part 3 of the queue restructure) from the already-capped list
    _compute_car_completion_stats() collected, plus the exact `remaining`
    count the caller already computed -- items_more is the gap between the
    two, never a second count of its own, so it can't drift out of sync."""
    items = stats.get(missing_items_key) or []
    return {
        "items": items,
        "items_more": max(0, remaining - len(items)),
    }


@app.route("/api/needs_category_queue")
@admin_required
def needs_category_queue():
    """Prioritized "Needs Category" work queue for the Train Matches
    dashboard. Tiers, in priority order:

      Tier 1 "Finish the gaps" -- fully_done_count > 0 AND < total_items:
        the car already has real progress on BOTH fronts, just isn't 100%
        done yet. Same underlying signal as needs_image_matching_queue()'s
        Tier 1.
      Tier 2 "Ready to tag" -- every item already has an image
        (image_linked_count == total_items) but zero categorization has
        started (category_set_count == 0): a clean, uninterrupted batch.
      Tier 3 -- everything else still missing at least one category (not
        started on images either, or some other partial state).

    Secondary sort within each tier: remaining (uncategorized) item count,
    descending, then car name ascending -- surfaces the biggest tagging
    opportunities first within a tier.

    Each car entry also carries "items" (up to QUEUE_ITEM_NAME_DISPLAY_CAP
    real stock item names still missing a category) and "items_more" (how
    many beyond that cap), so the dashboard can show specifics without a
    click-through (Part 3 of the queue restructure).
    """
    load_error = ensure_data_loaded()
    if load_error:
        return jsonify({"error": load_error}), 500

    stats_by_car = _compute_car_completion_stats()
    tier1, tier2, tier3 = [], [], []
    for car_model, stats in stats_by_car.items():
        total_items = stats["total_items"]
        category_set_count = stats["category_set_count"]
        if total_items <= 0 or category_set_count >= total_items:
            continue  # fully categorized (or empty car) -- not in this queue

        image_linked_count = stats["image_linked_count"]
        fully_done_count = stats["fully_done_count"]
        remaining = total_items - category_set_count
        entry = {
            "car": car_model,
            "total_items": total_items,
            "image_linked_count": image_linked_count,
            "category_set_count": category_set_count,
            "fully_done_count": fully_done_count,
            "remaining": remaining,
            **_queue_item_name_fields(stats, "missing_category_items", remaining),
        }

        if fully_done_count > 0 and fully_done_count < total_items:
            entry["tier"] = 1
            entry["tier_label"] = "Finish the gaps"
            tier1.append(entry)
        elif image_linked_count == total_items and category_set_count == 0:
            entry["tier"] = 2
            entry["tier_label"] = "Ready to tag"
            tier2.append(entry)
        else:
            entry["tier"] = 3
            entry["tier_label"] = "Needs category"
            tier3.append(entry)

    for bucket in (tier1, tier2, tier3):
        _sort_queue_bucket(bucket, "remaining")

    return jsonify({
        "cars": tier1 + tier2 + tier3,
        "secondary_sort": "remaining (uncategorized) item count, descending, then car name ascending",
    })


@app.route("/api/needs_image_matching_queue")
@admin_required
def needs_image_matching_queue():
    """Prioritized "Needs Image Matching" work queue. Tiers, in priority
    order:

      Tier 1 "Finish the gaps" -- the same fully_done_count > 0 AND <
        total_items signal as needs_category_queue()'s Tier 1 (shared across
        both queues).
      Tier 2 "Has category, needs image" (restored) -- every item already
        has a category (category_set_count == total_items) but the car
        still has zero fully-done items, i.e. zero images linked either.
        Previously deprioritized as rare; brought back as its own bucket
        between "finish the gaps" and the general remaining-items tier.
      Tier 3 -- everything else still missing at least one image (no
        category progress either, or some other partial state).

    Secondary sort within each tier: remaining (unmatched) item count,
    descending, then car name ascending.

    Each car entry also carries "items" (up to QUEUE_ITEM_NAME_DISPLAY_CAP
    real stock item names still missing an image) and "items_more" (how
    many beyond that cap), so the dashboard can show specifics without a
    click-through (Part 3 of the queue restructure).
    """
    load_error = ensure_data_loaded()
    if load_error:
        return jsonify({"error": load_error}), 500

    stats_by_car = _compute_car_completion_stats()
    tier1, tier2, tier3 = [], [], []
    for car_model, stats in stats_by_car.items():
        total_items = stats["total_items"]
        image_linked_count = stats["image_linked_count"]
        if total_items <= 0 or image_linked_count >= total_items:
            continue  # every item already has an image (or empty car)

        category_set_count = stats["category_set_count"]
        fully_done_count = stats["fully_done_count"]
        remaining = total_items - image_linked_count
        entry = {
            "car": car_model,
            "total_items": total_items,
            "image_linked_count": image_linked_count,
            "category_set_count": category_set_count,
            "fully_done_count": fully_done_count,
            "remaining": remaining,
            **_queue_item_name_fields(stats, "missing_image_items", remaining),
        }

        if fully_done_count > 0 and fully_done_count < total_items:
            entry["tier"] = 1
            entry["tier_label"] = "Finish the gaps"
            tier1.append(entry)
        elif category_set_count >= total_items:
            entry["tier"] = 2
            entry["tier_label"] = "Has category, needs image"
            tier2.append(entry)
        else:
            entry["tier"] = 3
            entry["tier_label"] = "Needs image matching"
            tier3.append(entry)

    for bucket in (tier1, tier2, tier3):
        _sort_queue_bucket(bucket, "remaining")

    return jsonify({
        "cars": tier1 + tier2 + tier3,
        "secondary_sort": "remaining (unmatched) item count, descending, then car name ascending",
    })


@app.route("/api/resolve_car_from_folder")
def resolve_car_from_folder():
    folder = (request.args.get("folder") or "").strip()
    if not folder:
        return jsonify({"car": None})
    # reuse internal hinting logic to map folder names to car models
    hint = _resolve_car_model_hint({"car_folder": folder})
    return jsonify({"car": hint})


@app.route("/admin/accounts")
@admin_required
@accounts_access_required(is_page=True)
def admin_accounts():
    status_message = (request.args.get("status") or "").strip()
    error_message = (request.args.get("error") or "").strip()
    return render_template(
        "accounts.html",
        status_message=status_message,
        error_message=error_message,
    )


@app.route("/admin/accounts/unlock", methods=["POST"])
@admin_required
def admin_accounts_unlock():
    if not _accounts_password_configured():
        return jsonify({
            "error": "Accounts panel is not configured. Set ACCOUNTS_ACCESS_PASSWORD in .env to enable it."
        }), 403

    next_url = _safe_next_url(request.form.get("next"))
    submitted_password = (request.form.get("password") or "").strip()

    rate_limit_key = f"accounts:{session.get('username', '')}"
    if not _check_login_rate_limit(rate_limit_key):
        return render_template(
            "accounts_unlock.html",
            error="Too many attempts. Please try again later.",
            next_url=next_url,
        ), 429

    if not submitted_password or submitted_password != Config.ACCOUNTS_ACCESS_PASSWORD:
        return render_template(
            "accounts_unlock.html",
            error="Incorrect password.",
            next_url=next_url,
        ), 401

    session["accounts_unlocked"] = True
    return redirect(next_url)


@app.route("/admin/create_user", methods=["POST"])
@admin_required
@accounts_access_required
def admin_create_user():
    payload = request.get_json(silent=True) or request.form or {}
    username = (payload.get("username") or "").strip()
    access_code = (payload.get("access_code") or "").strip()

    if not username:
        return jsonify({"success": False, "error": "username is required"}), 400
    if not access_code:
        return jsonify({"success": False, "error": "access code is required"}), 400

    try:
        created_user = db.create_customer_user(username, access_code)
    except ValueError as exc:
        message = str(exc)
        status = 409 if "exists" in message.lower() else 400
        return jsonify({"success": False, "error": message}), status

    db.log_account_action(created_user["id"], "created", _current_user_id())

    return jsonify({
        "success": True,
        "message": f"Account '{created_user['username']}' created successfully",
        "user": {
            "id": created_user["id"],
            "username": created_user["username"],
            "role": created_user.get("role", "customer"),
            "status": "active",
            "created_at": created_user.get("created_at"),
        },
    }), 200


@app.route("/admin/get_all_customers")
@admin_required
@accounts_access_required
def admin_get_all_customers():
    customers = db.get_all_customers_with_details()
    return jsonify([
        {
            "id": customer["id"],
            "username": customer["username"],
            "access_code": customer.get("access_code"),
            "status": "active" if customer.get("is_active", 1) else "paused",
            "is_active": bool(customer.get("is_active", 1)),
            "created_at": customer.get("created_at"),
            "last_login": customer.get("last_login"),
        }
        for customer in customers
    ])


@app.route("/admin/toggle_user_status/<int:user_id>", methods=["POST"])
@admin_required
@accounts_access_required
def admin_toggle_user_status(user_id):
    try:
        updated_user = db.toggle_customer_active_status(user_id)
    except ValueError as exc:
        message = str(exc)
        status = 404 if "not found" in message.lower() else 400
        return jsonify({"success": False, "error": message}), status

    is_active = bool(updated_user.get("is_active"))
    action = "resumed" if is_active else "paused"
    db.log_account_action(user_id, action, _current_user_id())
    return jsonify({
        "success": True,
        "message": f"Account '{updated_user['username']}' {action}",
        "user_id": user_id,
        "is_active": is_active,
    })


@app.route("/admin/set_all_customer_status", methods=["POST"])
@admin_required
@accounts_access_required
def admin_set_all_customer_status():
    payload = request.get_json(silent=True) or request.form or {}
    is_active = bool(payload.get("is_active", True))
    updated_count = db.set_all_customer_active_status(is_active)
    action = "resumed" if is_active else "paused"
    db.log_account_action(None, f"bulk_{action}", _current_user_id())
    return jsonify({
        "success": True,
        "message": f"{updated_count} account{'s' if updated_count != 1 else ''} {action}",
        "is_active": is_active,
        "updated_count": updated_count,
    })


@app.route("/admin/delete_user/<int:user_id>", methods=["POST"])
@admin_required
@accounts_access_required
def admin_delete_user(user_id):
    try:
        deleted_user = db.delete_customer_user(user_id)
    except ValueError as exc:
        message = str(exc)
        status = 404 if "not found" in message.lower() else 400
        return jsonify({"success": False, "error": message}), status

    db.log_account_action(user_id, "deleted", _current_user_id())
    return jsonify({
        "success": True,
        "message": f"Account '{deleted_user['username']}' deleted successfully",
        "user_id": user_id,
    })


@app.route("/train")
def train():
    load_error = ensure_data_loaded()
    if load_error:
        return jsonify({"error": load_error}), 500

    initial_image = None
    image_id = request.args.get("image_id", type=int)
    stock_item = request.args.get("stock_item", default="")
    car = request.args.get("car", default="")

    if image_id:
        initial_image = db.get_image_by_id(image_id)
    if initial_image is None:
        initial_image = db.get_next_unmapped_image()

    return render_template(
        "train.html",
        initial_image=initial_image,
        target_stock_item=stock_item,
        target_car=car,
        ss_image_folders=db.get_image_folders(),
        mapping_stats=db.get_mapping_stats(),
        category_stats=_compute_category_completion_stats(),
        selected_role=_current_role(),
        max_image_size=Config.MAX_IMAGE_SIZE,
        allowed_image_extensions=sorted(Config.ALLOWED_IMAGE_EXTENSIONS),
    )


@app.route("/bulk_match")
@admin_required
def bulk_match():
    return render_template("bulk_match.html")


@app.route("/api/search_all_stock_items")
@admin_required
def search_all_stock_items():
    load_error = ensure_data_loaded()
    if load_error:
        return jsonify({"error": load_error}), 500

    query = (request.args.get("q") or "").strip()
    type_filter = (request.args.get("type") or "").strip()
    color_filter = (request.args.get("color") or "").strip()

    items = _load_training_hierarchy_items()

    if type_filter or color_filter:
        type_norm = normalize_text(type_filter)
        color_norm = normalize_text(color_filter)
        matched = []
        for item in items:
            item_type, item_colors = extract_type_and_color(item.get("stock_item_name", ""))
            if type_norm and normalize_text(item_type or "") != type_norm:
                continue
            if color_norm and normalize_text("-".join(item_colors)) != color_norm:
                continue
            matched.append(item)
    else:
        normalized_query = normalize_text(query)
        matched = [
            item for item in items
            if normalized_query and normalized_query in normalize_text(item.get("stock_item_name", ""))
        ]

    seen = set()
    deduped = []
    for item in matched:
        key = (normalize_text(item.get("car_model", "")), normalize_text(item.get("stock_item_name", "")))
        if key in seen:
            continue
        seen.add(key)
        deduped.append(item)

    result_limit = 500
    truncated = len(deduped) > result_limit
    limited = deduped[:result_limit]

    # Reuses the same mapping lookup db.get_mappings_for_stock_items() that
    # _build_design_payload() uses, batched instead of one query per item —
    # same underlying query as db.get_mapping_for_stock_item() (used by
    # /get_current_mapping_image), just the many-items form of it.
    mapping_lookup = db.get_mappings_for_stock_items([item.get("stock_item_name", "") for item in limited])

    results = []
    for item in limited:
        stock_item_name = item.get("stock_item_name", "")
        mapping = mapping_lookup.get(_normalize_lookup_key(stock_item_name))
        results.append({
            "stock_item_name": stock_item_name,
            "car_model": item.get("car_model", ""),
            "is_mapped": bool(mapping and mapping.get("image_id")),
            "mapped_image_id": mapping.get("image_id") if mapping else None,
        })

    return jsonify({
        "results": results,
        "count": len(results),
        "total_matches": len(deduped),
        "truncated": truncated,
    })


@app.route("/api/list_product_categories")
@admin_required
def list_product_categories():
    load_error = ensure_data_loaded()
    if load_error:
        return jsonify({"error": load_error}), 500
    return jsonify(_get_product_categories())


@app.route("/admin/bulk_confirm_mapping", methods=["POST"])
@admin_required
def bulk_confirm_mapping():
    payload = request.get_json(silent=True) or {}
    try:
        image_id = int(payload.get("image_id"))
    except (TypeError, ValueError):
        return jsonify({"error": "image_id is required"}), 400

    stock_items = payload.get("stock_items")
    if not isinstance(stock_items, list) or not stock_items:
        return jsonify({"error": "stock_items must be a non-empty list"}), 400

    confirmed_count = 0
    failed = []
    stats = db.get_mapping_stats()

    for raw_stock_item in stock_items:
        stock_item_name = str(raw_stock_item or "").strip()
        if not stock_item_name:
            failed.append({"stock_item": raw_stock_item, "reason": "empty stock item name"})
            continue
        try:
            result = _confirm_mapping_core(image_id, stock_item_name, confirmed_by="human")
            confirmed_count += 1
            stats = result["stats"]
        except ValueError as exc:
            failed.append({"stock_item": stock_item_name, "reason": str(exc)})
        except Exception as exc:
            logger.exception("bulk_confirm_mapping failed for stock_item=%s", stock_item_name)
            failed.append({"stock_item": stock_item_name, "reason": str(exc)})

    return jsonify({
        "confirmed_count": confirmed_count,
        "failed": failed,
        "stats": stats,
    })


@app.route("/admin/assign_category", methods=["POST"])
@admin_required
def assign_category():
    """Bulk-assigns (last-write-wins, see design_categories table/Part 1) a
    material-tier category to every stock item name in the request. Unlike
    image-to-stock-item mapping, this isn't scoped to a single image_id --
    it's driven by the car-scoped "Assign Category" selection mode in
    index.html (see Part 2's /api/get_all_items_for_car), so many stock
    items are tagged in one call."""
    payload = request.get_json(silent=True) or {}
    category = str(payload.get("category") or "").strip()
    stock_item_names = payload.get("stock_item_names")

    if not db.category_exists(category):
        return jsonify({"error": f"category '{category}' does not exist"}), 400
    if not isinstance(stock_item_names, list) or not stock_item_names:
        return jsonify({"error": "stock_item_names must be a non-empty list"}), 400

    admin_user_id = _current_user_id()
    assigned_names = []
    failed = []
    for raw_name in stock_item_names:
        stock_item_name = str(raw_name or "").strip()
        if not stock_item_name:
            continue
        try:
            db.upsert_design_category(stock_item_name, category, admin_user_id)
            assigned_names.append(stock_item_name)
        except ValueError as exc:
            failed.append({"stock_item": stock_item_name, "reason": str(exc)})

    if assigned_names:
        _invalidate_queue_stats_cache()

    # There's no dedicated admin-action audit table for anything other than
    # user-account lifecycle events (account_logs is FK'd to a target
    # users.id row, which a stock item category assignment has no
    # equivalent of) -- logger.info() is the established pattern this
    # codebase already uses for other auditable admin/operational actions
    # (full refresh, stock updates; see logger.info calls elsewhere in this
    # file), so this follows that precedent instead of forcing a fit into
    # account_logs.
    logger.info(
        "Category '%s' assigned to %d item(s) by user_id=%s: %s",
        category, len(assigned_names), admin_user_id, assigned_names,
    )

    # Fire-and-forget pre-generation of the badged share variant (Part 7B)
    # for every just-tagged item that already has a mapped image, so the
    # cache is warm before anyone shares it -- runs on a background daemon
    # thread precisely so this response doesn't wait on however many images
    # need badging. Items with no mapped image yet have nothing to
    # pre-generate; get_share_image_badged() still builds on first request
    # either way (this is a pure optimization, not a correctness dependency).
    if assigned_names:
        mapping_lookup = db.get_mappings_for_stock_items(assigned_names)
        image_ids = [m["image_id"] for m in mapping_lookup.values() if m.get("image_id")]
        if image_ids:
            threading.Thread(target=_pregenerate_badged_images, args=(image_ids,), daemon=True).start()

    return jsonify({
        "assigned_count": len(assigned_names),
        "category": category,
        "failed": failed,
    })


@app.route("/api/categories")
def list_categories():
    """Live category list -- {name, abbreviation, sort_order} in display
    order. Read-only and not admin-gated (unlike every mutation route below):
    the on-thumbnail ribbon and category badge are visible to customers too
    (see _build_design_payload()'s comment), so this needs to be reachable by
    any logged-in role, not just admin. Used by index.html to rebuild the
    ribbon abbreviation map and the "Assign Category" picker immediately
    after a Category Settings change, without a full page reload (Part 5.7)."""
    return jsonify({"categories": db.get_all_categories()})


@app.route("/admin/categories/add", methods=["POST"])
@admin_required
def add_category_route():
    payload = request.get_json(silent=True) or {}
    name = str(payload.get("name") or "").strip()
    try:
        category = db.add_category(name)
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400

    _invalidate_queue_stats_cache()
    logger.info("Category '%s' added by user_id=%s", category["name"], _current_user_id())
    return jsonify({"category": category})


@app.route("/admin/categories/<string:name>/usage")
@admin_required
def category_usage(name):
    """Preview step (Part 3.3) -- how many items currently carry this
    category, so the frontend can show the real affected-item count in the
    delete warning before the admin confirms anything."""
    if not db.category_exists(name):
        return jsonify({"error": f"category '{name}' does not exist"}), 404
    return jsonify({"name": name, "affected_count": db.count_items_for_category(name)})


@app.route("/admin/categories/rename", methods=["POST"])
@admin_required
def rename_category_route():
    payload = request.get_json(silent=True) or {}
    old_name = str(payload.get("old_name") or "").strip()
    new_name = str(payload.get("new_name") or "").strip()

    if not db.category_exists(old_name):
        return jsonify({"error": f"category '{old_name}' does not exist"}), 404
    if not new_name:
        return jsonify({"error": "new_name is required"}), 400

    # A DIFFERENT existing category (case-insensitive) blocks a direct
    # rename -- the frontend needs to offer a merge instead (Part 3.2).
    # Renaming to the exact same name only with different casing (e.g. "Saka"
    # -> "SAKA") is not a conflict with a different category, so that still
    # goes through the normal rename path below.
    conflict = db.find_category_by_name_ci(new_name, exclude_name=old_name)
    if conflict:
        return jsonify({
            "conflict": True,
            "existing_name": conflict["name"],
            "affected_count": db.count_items_for_category(old_name),
        })

    try:
        affected = db.rename_category(old_name, new_name)
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400

    _invalidate_queue_stats_cache()
    logger.info(
        "Category '%s' renamed to '%s' by user_id=%s (%d item(s) affected)",
        old_name, new_name, _current_user_id(), affected,
    )
    if affected:
        _regenerate_badges_for_stock_items(_stock_item_names_for_category(new_name), category_name=new_name)

    return jsonify({"category": new_name, "affected_count": affected})


@app.route("/admin/categories/merge", methods=["POST"])
@admin_required
def merge_categories_route():
    payload = request.get_json(silent=True) or {}
    source_name = str(payload.get("source_name") or "").strip()
    target_name = str(payload.get("target_name") or "").strip()

    if not db.category_exists(source_name):
        return jsonify({"error": f"category '{source_name}' does not exist"}), 404
    if not db.category_exists(target_name):
        return jsonify({"error": f"category '{target_name}' does not exist"}), 404

    try:
        affected = db.merge_categories(source_name, target_name)
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400

    _invalidate_queue_stats_cache()
    logger.info(
        "Category '%s' merged into '%s' by user_id=%s (%d item(s) affected)",
        source_name, target_name, _current_user_id(), affected,
    )
    if affected:
        _regenerate_badges_for_stock_items(_stock_item_names_for_category(target_name), category_name=target_name)

    return jsonify({"category": target_name, "affected_count": affected})


@app.route("/admin/categories/delete", methods=["POST"])
@admin_required
def delete_category_route():
    payload = request.get_json(silent=True) or {}
    name = str(payload.get("name") or "").strip()

    if not db.category_exists(name):
        return jsonify({"error": f"category '{name}' does not exist"}), 404

    try:
        affected_names = db.delete_category(name)
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400

    _invalidate_queue_stats_cache()
    logger.info(
        "Category '%s' deleted by user_id=%s (%d item(s) reverted to uncategorized)",
        name, _current_user_id(), len(affected_names),
    )

    # The now-uncategorized items have nothing left to badge -- clean up
    # (not regenerate) their stale cached badge files synchronously (plain
    # filesystem deletes, fast enough not to need a background thread).
    if affected_names:
        mapping_lookup = db.get_mappings_for_stock_items(affected_names)
        for mapping in mapping_lookup.values():
            image_id = mapping.get("image_id")
            if image_id:
                _remove_badge_cache_for_image(image_id)

    return jsonify({"category": name, "affected_count": len(affected_names)})


def _stock_item_names_for_category(category_name):
    """All stock_item_name values currently tagged with category_name --
    used to scope badge regeneration to exactly the items a rename/merge/
    abbreviation-override just affected."""
    with db.get_connection() as conn:
        rows = conn.execute(
            "SELECT stock_item_name FROM design_categories WHERE category = ?", (category_name,)
        ).fetchall()
    return [row["stock_item_name"] for row in rows]


@app.route("/admin/categories/regen_status")
@admin_required
def category_badge_regen_status():
    """Polling endpoint for the "Updating N of M images..." indicator
    (see _BADGE_REGEN_STATUS above) -- purely informational. Badges are
    already guaranteed correct on every request regardless of this job's
    progress (see _get_category_info_for_image()/_ensure_badged_share_cached()),
    so `show` only gates whether a UI is worth displaying, never anything
    that blocks Share Images or any other action."""
    status = _badge_regen_status_snapshot()
    status["show"] = bool(status["running"] and status["total"] > _BADGE_REGEN_PROGRESS_THRESHOLD)
    return jsonify(status)


@app.route("/get_unmapped_images")
def get_unmapped_images_route():
    limit = _coerce_limit(request.args.get("limit", default=1, type=int), default_value=1)
    after = request.args.get("after", type=int)

    if after is not None:
        next_image = db.get_next_unmapped_image(after_image_id=after)
        images = [next_image] if next_image else []
    else:
        images = db.get_unmapped_images(limit=limit)

    return jsonify({
        "count": len(images),
        "first_image": images[0] if images else None,
        "images": images,
        "stats": db.get_mapping_stats(),
    })


@app.route("/train_images")
def train_images():
    # When a car folder is selected, return all images for that folder.
    # Without a folder filter, return only unmapped images (legacy behavior).
    folder = request.args.get("folder") or request.args.get("car")
    limit = _coerce_limit(request.args.get("limit", type=int), default_value=MAX_IMAGE_RESPONSE_LIMIT)

    try:
        if folder:
            images = db.get_images_by_folder(folder, limit=limit)
        else:
            images = db.get_unmapped_images(limit=limit if limit is not None else None)
    except Exception:
        # fallback to safe behavior
        images = db.get_unmapped_images(limit=limit if limit is not None else MAX_IMAGE_RESPONSE_LIMIT)

    return jsonify({
        "count": len(images),
        "images": images,
        "stats": db.get_mapping_stats(),
    })


@app.route("/export_images", methods=["POST"])
@admin_required
def export_images():
    payload = request.get_json(silent=True) or request.form or {}
    raw_image_ids = payload.get("image_ids") or []

    if isinstance(raw_image_ids, str):
        raw_image_ids = [item for item in re.split(r"[\s,]+", raw_image_ids) if item]

    image_ids = []
    for value in raw_image_ids:
        try:
            image_ids.append(int(value))
        except (TypeError, ValueError):
            continue

    if not image_ids:
        return jsonify({"error": "image_ids is required"}), 400

    archive_buffer = BytesIO()
    exported_count = 0

    with zipfile.ZipFile(archive_buffer, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        for image_id in image_ids:
            image_record = db.get_image_by_id(image_id)
            if not image_record:
                continue

            file_path = _resolve_stored_image_path(image_record.get("filepath"))
            if not file_path:
                continue

            car_folder = re.sub(r"[^A-Za-z0-9._ -]+", "_", str(image_record.get("car_folder") or "images")).strip() or "images"
            image_name = re.sub(r"[^A-Za-z0-9._ -]+", "_", str(image_record.get("filename") or f"image_{image_id}")).strip() or f"image_{image_id}"
            extension = os.path.splitext(file_path)[1] or ".jpg"
            archive_name = f"{car_folder}/{image_id}_{image_name}{extension}"
            archive.write(file_path, arcname=archive_name)
            exported_count += 1

    if exported_count == 0:
        return jsonify({"error": "No image files were available for export"}), 404

    archive_buffer.seek(0)
    return send_file(
        archive_buffer,
        mimetype="application/zip",
        as_attachment=True,
        download_name="ss-images.zip",
    )


@app.route("/remove_mapping", methods=["POST"])
@admin_required
def remove_mapping():
    payload = request.get_json(silent=True) or request.form or {}
    image_id = None
    try:
        image_id = int(payload.get("image_id"))
    except (TypeError, ValueError):
        image_id = None

    if not image_id:
        return jsonify({"error": "image_id is required"}), 400

    image_record = db.get_image_by_id(image_id)
    if not image_record:
        return jsonify({"error": "image not found"}), 404

    removed = db.remove_mapping_by_image_id(image_id)
    if removed:
        _invalidate_queue_stats_cache()
    return jsonify({
        "status": "removed" if removed else "no_mapping",
        "removed": bool(removed),
        "image_id": image_id,
        "stats": db.get_mapping_stats(),
    })


def _confirm_mapping_core(image_id, stock_item_name, car_model="", confidence=1.0, confirmed_by="human"):
    """Save image_id -> stock_item_name and return {image_id, next_image, stats}.

    Shared by the manual /confirm_mapping route and the image upload route so
    both go through the exact same save/overwrite behavior.
    """
    image_record = db.get_image_by_id(image_id)
    if not image_record:
        raise ValueError("image not found")

    resolved_car_model = car_model or _resolve_car_model_hint(image_record) or image_record.get("car_folder")
    if stock_item_name not in ("", "__UNMATCHABLE__"):
        db.remove_mappings_for_stock_item(stock_item_name, exclude_image_id=image_id)
    db.add_mapping(image_id, stock_item_name, resolved_car_model, confidence, confirmed_by=confirmed_by)
    if resolved_car_model and confidence >= 1.0 and stock_item_name not in ("", "__UNMATCHABLE__"):
        db.add_folder_mapping(image_record.get("car_folder", ""), resolved_car_model)
    _invalidate_queue_stats_cache()

    return {
        "image_id": image_id,
        "next_image": db.get_next_unmapped_image(after_image_id=image_id),
        "stats": db.get_mapping_stats(),
    }


def _apply_confirm_category_update(stock_item_name, category_raw):
    """Optional category set/clear alongside a confirm-match/upload-image
    save (Training Mode's category picker, additive to the standalone
    Assign Category flow). Reuses upsert_design_category()/
    db.remove_design_category() for all validation and writes -- this only
    decides whether a write is even needed and reuses the same badge-
    regen/queue-cache-invalidation triggers assign_category() already uses.

    Returns None if nothing needed doing: category_raw is None (field
    omitted from the request -- keeps old callers like bulk_confirm_mapping
    byte-for-byte unchanged) or it already matches the item's current
    category (skips the write entirely so assigned_at/assigned_by aren't
    bumped for a no-op, which upsert_design_category would otherwise do).
    Otherwise returns {"changed": bool, "category": str|None, "error": str|None}
    describing the outcome, so the caller can report partial success if the
    mapping saved but this failed.
    """
    if category_raw is None:
        return None

    category = str(category_raw).strip()
    current_category = db.get_category_for_stock_item(stock_item_name)
    if category == (current_category or ""):
        return None

    admin_user_id = _current_user_id()
    try:
        if category:
            db.upsert_design_category(stock_item_name, category, admin_user_id)
        else:
            db.remove_design_category(stock_item_name)
    except ValueError as exc:
        return {"changed": False, "category": current_category, "error": str(exc)}
    except Exception:
        logger.exception("Category update failed for stock_item_name=%s", stock_item_name)
        return {"changed": False, "category": current_category, "error": "failed to save category"}

    _invalidate_queue_stats_cache()
    _regenerate_badges_for_stock_items([stock_item_name], category_name=category or None)
    return {"changed": True, "category": category or None, "error": None}


@app.route("/confirm_mapping", methods=["POST"])
@admin_required
def confirm_mapping():
    payload = request.get_json(silent=True) or request.form or {}
    image_id = None
    try:
        image_id = int(payload.get("image_id"))
    except Exception:
        image_id = None
    if not image_id:
        return jsonify({"error": "image_id is required"}), 400

    stock_item_name = payload.get("stock_item_name", "")
    confidence = payload.get("confidence", 1.0)
    confirmed_by = payload.get("confirmed_by", "human")
    car_model = payload.get("car_model", "")

    try:
        confidence = float(confidence)
    except (TypeError, ValueError):
        confidence = 1.0

    try:
        result = _confirm_mapping_core(image_id, stock_item_name, car_model=car_model, confidence=confidence, confirmed_by=confirmed_by)
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 404

    response = {
        "status": "saved",
        "image_id": result["image_id"],
        "next_image": result["next_image"],
        "stats": result["stats"],
    }

    # Optional category set/clear in the same request (Training Mode's
    # category picker) -- "category" only appears in the payload at all when
    # the picker sent one, so a caller that never mentions it (e.g.
    # bulk_confirm_mapping's own JSON body) leaves categories untouched,
    # identical to before this existed.
    if "category" in payload and stock_item_name not in ("", "__UNMATCHABLE__"):
        category_update = _apply_confirm_category_update(stock_item_name, payload.get("category"))
        if category_update is not None:
            response["category_update"] = category_update
            if category_update.get("error"):
                response["status"] = "saved_category_failed"
            if category_update.get("changed"):
                # Lets the dashboard's category-completion stat card (Part 4
                # of the queue restructure) refresh itself in place, the
                # same way updateProgress() already does for mapping_stats.
                response["category_stats"] = _compute_category_completion_stats()

    return jsonify(response)


def _sanitize_upload_car_folder(raw_name):
    value = str(raw_name or "").strip()
    if not value:
        raise ValueError("Car folder name is required")
    if "\x00" in value:
        raise ValueError("Car folder name contains invalid characters")
    if "/" in value or "\\" in value or ".." in value:
        raise ValueError("Car folder name cannot contain path separators")
    return value


def _sanitize_upload_filename(original_name):
    base = os.path.basename(str(original_name or "").replace("\\", "/")).strip()
    return base or "upload"


def _unique_filename_in_dir(target_dir, base_name):
    name_part, ext_part = os.path.splitext(base_name)
    candidate = base_name
    counter = 1
    while os.path.exists(os.path.join(target_dir, candidate)):
        counter += 1
        candidate = f"{name_part} ({counter}){ext_part}"
    return candidate


@app.route("/admin/upload_image", methods=["POST"])
@admin_required
def admin_upload_image():
    uploaded_file = request.files.get("file")
    stock_item_name = (request.form.get("stock_item") or "").strip()

    if not uploaded_file or not uploaded_file.filename:
        return jsonify({"success": False, "error": "No file was uploaded"}), 400
    if not stock_item_name:
        return jsonify({"success": False, "error": "stock_item is required"}), 400

    try:
        car_folder = _sanitize_upload_car_folder(request.form.get("car_folder", ""))
    except ValueError as exc:
        return jsonify({"success": False, "error": str(exc)}), 400

    original_name = uploaded_file.filename
    extension = os.path.splitext(original_name)[1].lower()
    if extension not in Config.ALLOWED_IMAGE_EXTENSIONS:
        allowed = ", ".join(sorted(Config.ALLOWED_IMAGE_EXTENSIONS))
        return jsonify({"success": False, "error": f"Unsupported file type '{extension or 'unknown'}'. Allowed types: {allowed}"}), 400

    uploaded_file.seek(0, os.SEEK_END)
    file_size = uploaded_file.tell()
    uploaded_file.seek(0)
    if file_size <= 0:
        return jsonify({"success": False, "error": "Uploaded file is empty"}), 400
    if file_size > Config.MAX_IMAGE_SIZE:
        max_mb = Config.MAX_IMAGE_SIZE / (1024 * 1024)
        return jsonify({"success": False, "error": f"File is too large. Maximum size is {max_mb:.0f} MB"}), 400

    target_dir = os.path.join(IMAGE_SCAN_ROOT, car_folder)
    try:
        os.makedirs(target_dir, exist_ok=True)
    except OSError as exc:
        return jsonify({"success": False, "error": f"Could not create folder: {exc}"}), 500

    safe_base_name = _sanitize_upload_filename(original_name)
    final_filename = _unique_filename_in_dir(target_dir, safe_base_name)
    destination_path = os.path.join(target_dir, final_filename)

    try:
        uploaded_file.save(destination_path)
    except OSError as exc:
        return jsonify({"success": False, "error": f"Failed to save file: {exc}"}), 500

    relative_path = os.path.relpath(destination_path, IMAGE_SCAN_ROOT).replace("\\", "/")

    try:
        image_id = db.add_image(car_folder, final_filename, relative_path)
    except Exception as exc:
        try:
            os.remove(destination_path)
        except OSError:
            pass
        logger.exception("Failed to add uploaded image to database")
        return jsonify({"success": False, "error": f"Failed to save image record: {exc}"}), 500

    try:
        confirm_result = _confirm_mapping_core(image_id, stock_item_name, car_model=car_folder, confidence=1.0, confirmed_by="human")
    except ValueError as exc:
        logger.exception("Failed to confirm mapping for uploaded image")
        return jsonify({"success": False, "error": f"Image saved but failed to confirm match: {exc}"}), 500

    response_payload = {
        "success": True,
        "image_id": image_id,
        "image_url": f"/get_image/{image_id}",
        "car_folder": car_folder,
        "stock_item": stock_item_name,
        "mapped": True,
        "stats": confirm_result.get("stats"),
    }

    # Same optional category set/clear as /confirm_mapping (Training Mode's
    # Upload Image modal gets the identical picker) -- "category" is a
    # regular multipart form field here rather than JSON.
    if "category" in request.form:
        category_update = _apply_confirm_category_update(stock_item_name, request.form.get("category"))
        if category_update is not None:
            response_payload["category_update"] = category_update
            if category_update.get("error"):
                response_payload["success_partial"] = "category_failed"
            if category_update.get("changed"):
                response_payload["category_stats"] = _compute_category_completion_stats()

    return jsonify(response_payload), 200


def _resolve_stored_image_path(stored_path):
    if not stored_path:
        return None

    value = str(stored_path).strip().replace("\\", "/")
    if not value:
        return None

    # Preferred portable root: always resolve from app root.
    portable_root = os.path.join(app.root_path, "data", "S.S IMAGE")
    fallback_root = IMAGE_SCAN_ROOT

    def _safe_join(root_dir, relative_path):
        candidate = os.path.normpath(os.path.join(root_dir, relative_path))
        root_norm = os.path.normcase(os.path.normpath(root_dir)) if os.name == "nt" else os.path.normpath(root_dir)
        candidate_norm = os.path.normcase(candidate) if os.name == "nt" else candidate
        if candidate_norm == root_norm or candidate_norm.startswith(root_norm + os.sep):
            return candidate
        return None

    # Handle legacy absolute rows from older scans.
    if os.path.isabs(value):
        if os.path.exists(value):
            return value
        # Salvage by using just the relative suffix from ".../S.S IMAGE/" onward.
        marker = "s.s image/"
        lowered = value.lower().replace("\\", "/")
        marker_index = lowered.find(marker)
        if marker_index != -1:
            value = value[marker_index + len(marker):].replace("\\", "/")
        else:
            value = os.path.basename(value)

    # Strip optional root prefix if present in stored relative value.
    normalized = value.replace("\\", "/").lstrip("/")
    if normalized.lower().startswith("s.s image/"):
        normalized = normalized.split("/", 1)[1] if "/" in normalized else ""
    if not normalized:
        return None

    for root_dir in (portable_root, fallback_root):
        candidate = _safe_join(root_dir, normalized)
        if candidate and os.path.exists(candidate):
            return candidate
    return None


@app.route("/get_image/<int:image_id>")
def get_image(image_id):
    image_record = db.get_image_by_id(image_id)
    if not image_record:
        return _placeholder_response()

    file_path = _resolve_stored_image_path(image_record.get("filepath"))
    if file_path:
        return send_file(file_path, conditional=True)

    return _placeholder_response()


@app.route("/get_stock_image")
def get_stock_image():
    stock_item_name = request.args.get("stock_item", "")
    if not stock_item_name:
        return _placeholder_response()

    mapping = db.get_mapping_for_stock_item(stock_item_name)
    if mapping:
        file_path = _resolve_stored_image_path(mapping.get("filepath"))
        if file_path:
            return send_file(file_path, conditional=True)

    return _placeholder_response()


def _prepare_share_image(source_path):
    """Open, EXIF-transpose, flatten-to-RGB, and downscale source_path to
    SHARE_IMAGE_MAX_DIMENSION -- the resize/compress baseline shared by both
    the plain share image (_build_share_image) and the badged variant
    (_build_badged_share_image), so there's exactly one place that logic
    lives. Returns a standalone PIL Image (detached from the source file
    handle)."""
    with Image.open(source_path) as original:
        image = ImageOps.exif_transpose(original)  # bake in phone-camera rotation

        if image.mode in ("RGBA", "LA") or (image.mode == "P" and "transparency" in image.info):
            flattened = Image.new("RGB", image.size, (255, 255, 255))
            flattened.paste(image, mask=image.convert("RGBA").split()[-1])
            image = flattened
        elif image.mode != "RGB":
            image = image.convert("RGB")

        width, height = image.size
        scale = min(1.0, SHARE_IMAGE_MAX_DIMENSION / max(width, height))
        if scale < 1.0:
            image = image.resize(
                (max(1, round(width * scale)), max(1, round(height * scale))),
                Image.LANCZOS,
            )
        return image.copy()  # detach from the `with Image.open(...)` block


def _build_share_image(source_path, cache_path):
    """Write a resized, compressed JPEG copy of source_path to cache_path
    for the bulk share flow. If re-encoding doesn't actually save space (a
    handful of already-small, already-compressed originals re-encode
    slightly larger), the original bytes are cached under the same name
    unchanged instead -- callers never need to know which happened."""
    image = _prepare_share_image(source_path)
    buffer = BytesIO()
    image.save(buffer, format="JPEG", quality=SHARE_IMAGE_JPEG_QUALITY, optimize=True)
    compressed_bytes = buffer.getvalue()

    if len(compressed_bytes) >= os.path.getsize(source_path):
        shutil.copyfile(source_path, cache_path)
    else:
        with open(cache_path, "wb") as handle:
            handle.write(compressed_bytes)


_BADGE_FONT_CACHE = {}
# Common bold-ish TrueType font locations, tried in order. Windows ships
# arialbd.ttf/arial.ttf on essentially every install (this app targets
# Windows deployment -- see launcher.pyw/first_time_setup.bat); the Linux
# paths are a best-effort fallback for dev/CI environments. If none exist,
# Pillow's own bundled scalable default font (load_default(size=...), added
# in Pillow 10.1) is used instead -- there is always a usable fallback.
_BADGE_FONT_CANDIDATES = (
    "C:/Windows/Fonts/arialbd.ttf",
    "C:/Windows/Fonts/arial.ttf",
    "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf",
    "/usr/share/fonts/truetype/liberation/LiberationSans-Bold.ttf",
)


def _load_badge_font(size):
    if size in _BADGE_FONT_CACHE:
        return _BADGE_FONT_CACHE[size]

    from PIL import ImageFont

    font = None
    for candidate in _BADGE_FONT_CANDIDATES:
        if os.path.exists(candidate):
            try:
                font = ImageFont.truetype(candidate, size)
                break
            except Exception:
                continue
    if font is None:
        try:
            font = ImageFont.load_default(size=size)
        except TypeError:
            # Pillow < 10.1 doesn't accept a size kwarg here.
            font = ImageFont.load_default()

    _BADGE_FONT_CACHE[size] = font
    return font


_BADGE_MAX_BANNER_WIDTH_FRACTION = 0.35
_BADGE_PAD_X_RATIO = 0.6
_BADGE_PAD_Y_RATIO = 0.4


def _draw_category_badge(image, category_text):
    """Draws a small semi-transparent banner in the bottom-left corner of
    `image` with `category_text`, legible over varying image content
    underneath (a flat alpha-blended rectangle behind solid white text,
    rather than text alone, keeps it readable on both light and dark
    photos) -- including at the small (~200-300px wide) thumbnail size
    WhatsApp actually displays a shared image at before it's tapped to
    expand, which is a much harsher legibility test than viewing the
    full-resolution share image. Returns a new RGB image; `image` itself is
    not mutated."""
    from PIL import ImageDraw

    base = image.convert("RGBA")
    overlay = Image.new("RGBA", base.size, (0, 0, 0, 0))
    draw = ImageDraw.Draw(overlay)

    text = str(category_text or "").strip().upper()
    font_size = max(16, round(min(base.size) * 0.06))
    font = _load_badge_font(font_size)
    bbox = draw.textbbox((0, 0), text, font=font)
    text_w = bbox[2] - bbox[0]

    # The font size above scales off the image's shorter dimension so the
    # badge reads consistently across portrait/landscape photos -- but the
    # longer category names ("Pearl Designer", "Napa Designer") at that same
    # size can produce a banner wide enough to swallow most of a narrow
    # product photo, and on a few real catalog images that already carry
    # their own baked-in caption text near the bottom, wide enough to
    # visually collide with it. If the initial size would exceed
    # _BADGE_MAX_BANNER_WIDTH_FRACTION of the image width, the font shrinks
    # just enough to fit (one re-measure at the smaller size -- font metrics
    # aren't perfectly linear, but one pass is close enough for a visual
    # sizing fix and avoids unbounded looping) rather than always using the
    # same size regardless of text length.
    max_banner_w = base.size[0] * _BADGE_MAX_BANNER_WIDTH_FRACTION
    initial_banner_w = text_w + 2 * round(font_size * _BADGE_PAD_X_RATIO)
    if text_w > 0 and initial_banner_w > max_banner_w:
        font_size = max(12, round(font_size * (max_banner_w / initial_banner_w)))
        font = _load_badge_font(font_size)
        bbox = draw.textbbox((0, 0), text, font=font)
        text_w = bbox[2] - bbox[0]

    text_h = bbox[3] - bbox[1]
    pad_x, pad_y = round(font_size * _BADGE_PAD_X_RATIO), round(font_size * _BADGE_PAD_Y_RATIO)
    margin = max(10, round(min(base.size) * 0.025))

    banner_w = text_w + pad_x * 2
    banner_h = text_h + pad_y * 2
    x0 = margin
    y0 = base.size[1] - margin - banner_h
    x1 = x0 + banner_w
    y1 = y0 + banner_h

    # Alpha bumped from an earlier, lighter fill to 190/255 (~75% opaque) --
    # legible at full resolution either way, but a lighter fill visibly
    # washed out and lost contrast against busy/light image content once
    # JPEG-compressed and scaled down to a small chat thumbnail.
    draw.rectangle([x0, y0, x1, y1], fill=(0, 0, 0, 190))
    draw.text((x0 + pad_x - bbox[0], y0 + pad_y - bbox[1]), text, font=font, fill=(255, 255, 255, 255))

    return Image.alpha_composite(base, overlay).convert("RGB")


def _build_badged_share_image(source_path, cache_path, category):
    """Same resize/compress baseline as _build_share_image(), plus the
    category banner burned onto the resized image before final JPEG
    encoding. Unlike _build_share_image(), there's no "keep the original
    bytes if smaller" fallback -- the pixels are always modified, so the
    original file is never a valid substitute."""
    image = _prepare_share_image(source_path)
    badged = _draw_category_badge(image, category)
    buffer = BytesIO()
    badged.save(buffer, format="JPEG", quality=SHARE_IMAGE_JPEG_QUALITY, optimize=True)
    with open(cache_path, "wb") as handle:
        handle.write(buffer.getvalue())


def _category_slug(category):
    return re.sub(r"[^a-z0-9]+", "-", str(category or "").lower()).strip("-") or "category"


def _badged_share_cache_path(image_id, category):
    return os.path.join(SHARE_IMAGE_CACHE_DIR, f"{image_id}_badge_{_category_slug(category)}.jpg")


def _cleanup_stale_badge_variants(image_id, current_category):
    """Removes any previously-cached badge variant(s) for image_id whose
    filename doesn't match the current category -- the category is part of
    the cache filename (see _badged_share_cache_path()), so a re-tag alone
    would otherwise leave the old category's file behind as permanent,
    never-served disk clutter."""
    current_path = _badged_share_cache_path(image_id, current_category)
    pattern = os.path.join(SHARE_IMAGE_CACHE_DIR, f"{image_id}_badge_*.jpg")
    for stale_path in glob.glob(pattern):
        if os.path.normcase(stale_path) != os.path.normcase(current_path):
            try:
                os.remove(stale_path)
            except OSError:
                pass


def _get_category_info_for_image(image_id):
    """Returns {"name": ..., "abbreviation": ...} for image_id's assigned
    category, or None if it has no mapped stock item or no category. The
    full name is what gets burned onto the badged share image (see
    _ensure_badged_share_cached() below) -- the width-cap/font-shrink
    legibility logic in _draw_category_badge() handles long full names
    (it was originally built and tested against them, before the
    abbreviation field existed). The abbreviation is only for the
    on-thumbnail .category-ribbon (client-side, see index.html) and is not
    used here."""
    mapping = db.get_mapping_by_image_id(image_id)
    if not mapping or not mapping.get("stock_item_name"):
        return None
    category_name = db.get_category_for_stock_item(mapping["stock_item_name"])
    if not category_name:
        return None
    category_row = db.get_category_by_name(category_name)
    abbreviation = category_row["abbreviation"] if category_row else category_name
    return {"name": category_name, "abbreviation": abbreviation}


def _ensure_badged_share_cached(image_id, force=False):
    """Builds (if missing/stale/forced) and returns the cache path for
    image_id's badged share variant, or None if there's nothing to badge (no
    mapped image, unresolvable source file, or no category assigned). This is
    the ONE code path that builds this cache -- get_share_image_badged()
    (on-demand, from the share flow) and every fire-and-forget pre-generation/
    regeneration trigger (assign_category(), and Part 4/6's rename/merge/
    abbreviation-override paths via _regenerate_badges_for_stock_items())
    call this exact function, so there is no duplicate image-generation logic
    to drift out of sync.

    force=True bypasses the mtime staleness check and always rebuilds --
    needed when the category's NAME or its abbreviation text changed but the
    source photo itself didn't, which the passive mtime comparison alone
    would never notice (Part 4/6's regeneration is otherwise always
    forced=True)."""
    image_record = db.get_image_by_id(image_id)
    if not image_record:
        return None
    source_path = _resolve_stored_image_path(image_record.get("filepath"))
    if not source_path:
        return None
    category_info = _get_category_info_for_image(image_id)
    if not category_info:
        return None

    os.makedirs(SHARE_IMAGE_CACHE_DIR, exist_ok=True)
    cache_path = _badged_share_cache_path(image_id, category_info["name"])
    stale = force or not os.path.exists(cache_path) or os.path.getmtime(cache_path) < os.path.getmtime(source_path)
    if stale:
        tmp_path = f"{cache_path}.tmp{os.getpid()}_{image_id}"
        try:
            _build_badged_share_image(source_path, tmp_path, category_info["name"])
            os.replace(tmp_path, cache_path)
        except Exception:
            try:
                os.remove(tmp_path)
            except OSError:
                pass
            raise
        _cleanup_stale_badge_variants(image_id, category_info["name"])
    return cache_path


def _remove_badge_cache_for_image(image_id):
    """Deletes every cached badge variant for image_id outright (Part 3.3) --
    used when an item is deleted back to uncategorized, where there is no
    longer any category to regenerate a badge FOR, so the stale cached file
    must simply be cleaned up rather than rebuilt."""
    pattern = os.path.join(SHARE_IMAGE_CACHE_DIR, f"{image_id}_badge_*.jpg")
    for stale_path in glob.glob(pattern):
        try:
            os.remove(stale_path)
        except OSError:
            pass


# Progress visibility for bulk badge regeneration (rename/merge/abbreviation-
# override) -- NOT a correctness mechanism (see _get_category_info_for_image()/
# _ensure_badged_share_cached(): every request already derives its cache key
# and burned text from a live DB lookup, so a request during an in-flight
# regeneration job is always served fresh/correct, never stale -- confirmed
# by a real concurrent-request test, not just this comment). This is purely
# so an admin isn't left wondering whether a bulk change has "taken" yet.
# Same lightweight global-status-dict + polling-endpoint shape as
# full_refresh_status/_run_full_refresh_job() above -- counters accumulate
# additively across overlapping jobs (e.g. a rename immediately followed by
# an abbreviation override) rather than one job's start clobbering another's
# in-flight progress.
_BADGE_REGEN_PROGRESS_THRESHOLD = 3  # only worth surfacing a UI for real bulk jobs
_BADGE_REGEN_LOCK = threading.Lock()
_BADGE_REGEN_STATUS = {"running": False, "total": 0, "completed": 0, "category": None}


def _badge_regen_status_snapshot():
    with _BADGE_REGEN_LOCK:
        return dict(_BADGE_REGEN_STATUS)


def _start_badge_regen_job(count, category_name=None):
    if count <= 0:
        return
    with _BADGE_REGEN_LOCK:
        if not _BADGE_REGEN_STATUS["running"]:
            _BADGE_REGEN_STATUS["total"] = 0
            _BADGE_REGEN_STATUS["completed"] = 0
        _BADGE_REGEN_STATUS["total"] += count
        _BADGE_REGEN_STATUS["running"] = True
        _BADGE_REGEN_STATUS["category"] = category_name


def _mark_badge_regen_progress():
    with _BADGE_REGEN_LOCK:
        _BADGE_REGEN_STATUS["completed"] += 1
        if _BADGE_REGEN_STATUS["completed"] >= _BADGE_REGEN_STATUS["total"]:
            _BADGE_REGEN_STATUS["running"] = False


def _regenerate_badges_for_stock_items(stock_item_names, category_name=None):
    """Reusable trigger (Part 4) for forcing the badged share cache to
    rebuild for every mapped image behind the given stock item names --
    extracted from the single-item pre-generation assign_category() already
    used, so a rename/merge (Part 3.2) or an abbreviation override (Part 6.3)
    affecting many items at once shares the exact same generation logic.
    Fire-and-forget on a background daemon thread, same as the existing
    per-item trigger, so the admin's request never waits on however many
    images need re-badging. `category_name` is display-only context for the
    progress status (see _BADGE_REGEN_STATUS above)."""
    if not stock_item_names:
        return
    mapping_lookup = db.get_mappings_for_stock_items(stock_item_names)
    image_ids = [m["image_id"] for m in mapping_lookup.values() if m.get("image_id")]
    if image_ids:
        _start_badge_regen_job(len(image_ids), category_name)
        threading.Thread(target=_force_regenerate_badged_images, args=(image_ids,), daemon=True).start()


def _force_regenerate_badged_images(image_ids):
    for image_id in image_ids:
        try:
            _ensure_badged_share_cached(image_id, force=True)
        except Exception:
            logger.exception("Background badge regeneration failed for image_id=%s", image_id)
        finally:
            _mark_badge_regen_progress()


def _pregenerate_badged_images(image_ids):
    """Fire-and-forget background pre-warm of the badged share cache (see
    Part 3's assign_category() route, which starts this on a daemon thread
    so the admin's request never waits on it). Best-effort: a single
    image's failure is logged and skipped, never raised, since nothing is
    awaiting this thread's result -- get_share_image_badged() still
    regenerates on demand as a correctness fallback either way."""
    for image_id in image_ids:
        try:
            _ensure_badged_share_cached(image_id)
        except Exception:
            logger.exception("Background badge pre-generation failed for image_id=%s", image_id)


@app.route("/get_share_image/<int:image_id>")
def get_share_image(image_id):
    image_record = db.get_image_by_id(image_id)
    if not image_record:
        return _placeholder_response()

    source_path = _resolve_stored_image_path(image_record.get("filepath"))
    if not source_path:
        return _placeholder_response()

    os.makedirs(SHARE_IMAGE_CACHE_DIR, exist_ok=True)
    cache_path = os.path.join(SHARE_IMAGE_CACHE_DIR, f"{image_id}.jpg")

    # Regenerated only the first time (or if the source image was replaced
    # since -- re-scanning/re-mapping can swap in a different file under the
    # same image_id), not on every request.
    stale = not os.path.exists(cache_path) or os.path.getmtime(cache_path) < os.path.getmtime(source_path)
    if stale:
        tmp_path = f"{cache_path}.tmp{os.getpid()}"
        try:
            _build_share_image(source_path, tmp_path)
            os.replace(tmp_path, cache_path)
        except Exception:
            logger.exception("Failed to build share-optimized image for id %s; serving original", image_id)
            try:
                os.remove(tmp_path)
            except OSError:
                pass
            return send_file(source_path, conditional=True)

    return send_file(cache_path, conditional=True, mimetype="image/jpeg")


@app.route("/get_share_image_badged/<int:image_id>")
def get_share_image_badged(image_id):
    """Share-optimized image with the item's category burned onto it (see
    _draw_category_badge()). This is the single endpoint the frontend's
    "Share Images" flow now uses for every share, categorized or not --
    falls back to the exact same plain output as /get_share_image/<id>
    when the item has no category assigned, so it's always a safe drop-in
    regardless of whether the selected images are categorized.

    /get_share_image/<id> itself (get_share_image() below) is no longer
    called directly by any frontend flow as of the caption-sharing removal,
    but is NOT dead code -- it's still the fallback this function calls
    for uncategorized items, and its own route is left in place."""
    try:
        cache_path = _ensure_badged_share_cached(image_id)
    except Exception:
        logger.exception("Failed to build badged share image for id %s; falling back to plain share image", image_id)
        cache_path = None

    if cache_path:
        return send_file(cache_path, conditional=True, mimetype="image/jpeg")
    return get_share_image(image_id)


@app.route("/get_current_mapping_image")
def get_current_mapping_image():
    stock_item_name = request.args.get("stock_item", "")
    if not stock_item_name:
        return jsonify({"has_mapping": False, "image_id": None, "image_url": None, "confidence": None, "category": None})

    # category is looked up regardless of has_mapping -- Training Mode's
    # category picker (and the current-match preview's category label) need
    # to know the item's category even when it has no image yet.
    category = db.get_category_for_stock_item(stock_item_name)

    mapping = db.get_mapping_for_stock_item(stock_item_name)
    if not mapping:
        return jsonify({"has_mapping": False, "image_id": None, "image_url": None, "confidence": None, "category": category})

    image_id = mapping.get("image_id")
    return jsonify({
        "has_mapping": True,
        "image_id": image_id,
        "image_url": f"/get_image/{image_id}",
        "confidence": mapping.get("confidence"),
        "category": category,
    })


@app.route("/suggest_match/<int:image_id>")
def suggest_match_route(image_id):
    return jsonify({"error": "AI suggestion is disabled for now"}), 410


@app.route("/scan_images", methods=["POST"])
@admin_required
def scan_images():
    result = image_scanner.scan_ss_image_folder(IMAGE_SCAN_ROOT)
    missing = image_scanner.find_missing_image_rows(IMAGE_SCAN_ROOT)
    return jsonify({
        "status": "scanned",
        **result,
        "missing_count": missing["missing_count"],
        "missing_mapped_count": missing["missing_mapped_count"],
        "over_threshold_warning": missing["over_threshold_warning"],
        "missing_image_ids": [row["id"] for row in missing["rows"]],
        "missing_images": [
            {
                "id": row["id"],
                "car_folder": row.get("car_folder"),
                "filename": row.get("filename"),
                "filepath": row.get("filepath"),
                "mapped": bool(row.get("mapped")),
                "stock_item_names": row.get("stock_item_names") or [],
            }
            for row in missing["rows"]
        ],
        "stats": db.get_mapping_stats(),
    })


@app.route("/admin/remove_missing_images", methods=["POST"])
@admin_required
def remove_missing_images():
    payload = request.get_json(silent=True) or {}
    raw_ids = payload.get("image_ids")
    if not isinstance(raw_ids, list) or not raw_ids:
        return jsonify({"error": "image_ids is required"}), 400

    try:
        requested_ids = {int(image_id) for image_id in raw_ids}
    except (TypeError, ValueError):
        return jsonify({"error": "image_ids must be a list of integers"}), 400

    # Re-verify from scratch rather than trusting the earlier scan result --
    # a file may have reappeared (drive reconnected, folder restored)
    # between the scan and this confirm click.
    missing = image_scanner.find_missing_image_rows(IMAGE_SCAN_ROOT)
    still_missing_ids = {row["id"] for row in missing["rows"]} & requested_ids

    images_removed, mappings_removed = db.remove_missing_image_rows(still_missing_ids)
    if mappings_removed:
        _invalidate_queue_stats_cache()

    return jsonify({
        "status": "removed",
        "images_removed": images_removed,
        "mappings_removed": mappings_removed,
        "requested_count": len(requested_ids),
        "skipped_count": len(requested_ids) - len(still_missing_ids),
        "stats": db.get_mapping_stats(),
    })


@app.route("/mapping_stats")
def mapping_stats():
    return jsonify(db.get_mapping_stats())


def run_full_refresh_job():
    """Fetch car master, main hierarchy, and item stock from Tally, then reload.

    Shared by the manual /full_refresh route and the automatic startup refresh.
    Caller is responsible for holding FULL_REFRESH_LOCK for the duration of this call.
    """
    global full_refresh_status
    try:
        full_refresh_status = {
            "running": True,
            "stage": "car_master",
            "error": None,
            "timestamp": datetime.now().isoformat()
        }

        car_names = fetch_car_master_from_tally()
        save_car_master_to_file(car_names)
        full_refresh_status["stage"] = "main_hierarchy"

        flat_rows = fetch_main_hierarchy_from_tally()
        save_main_hierarchy_to_file(flat_rows)
        full_refresh_status["stage"] = "item_stock"

        fetch_item_stock_flat()
        full_refresh_status["stage"] = "reloading"

        load_data(refresh_first=False)
        full_refresh_status = {
            "running": False,
            "stage": "done",
            "error": None,
            "timestamp": datetime.now().isoformat()
        }
        logger.info("Full refresh completed successfully")

    except Exception as exc:
        error_code = _classify_tally_exception(exc)
        full_refresh_status = {
            "running": False,
            "stage": "error",
            "error": str(exc),
            "error_code": error_code,
            "timestamp": datetime.now().isoformat()
        }
        logger.exception("Full refresh failed: %s", exc)


@app.route("/full_refresh", methods=["POST"])
@admin_required
def full_refresh():
    if not FULL_REFRESH_LOCK.acquire(blocking=False):
        return jsonify({
            "ok": False,
            "busy": True,
            "message": "Full refresh already in progress"
        }), 409

    def _run_full_refresh():
        try:
            run_full_refresh_job()
        finally:
            FULL_REFRESH_LOCK.release()

    thread = threading.Thread(target=_run_full_refresh, daemon=True)
    thread.start()

    return jsonify({
        "ok": True,
        "busy": False,
        "message": "Full refresh started"
    }), 202


@app.route("/full_refresh_status")
def full_refresh_status_route():
    return jsonify(full_refresh_status)


@app.route("/reload", methods=["POST"])
@admin_required
def reload_data():
    """Reload cached data (without reaching out to Tally)."""
    try:
        load_data(refresh_first=False)
    except Exception as exc:
        load_error = str(exc)
        return jsonify({"status": "error", "error": load_error}), 500
    return jsonify({"status": "reloaded"})


@app.route("/update_stock", methods=["POST"])
@admin_required
def update_stock():
    return refresh_stock()

@app.route("/refresh_item_stock", methods=["POST"])
@admin_required
def refresh_item_stock():
    result = _refresh_stock_data()
    status_code = 200
    if not result.get("tally_online"):
        result["status"] = "using_last_saved_upload"
    return jsonify(result), status_code

@app.route("/last_update")
def last_update():
    """Return the last modification time of the item stock export file."""
    try:
        update_file = get_latest_stock_file_path() or ITEM_STOCK_FILE_AUTO
        if os.path.exists(update_file):
            mtime = os.path.getmtime(update_file)
            dt = datetime.fromtimestamp(mtime)
            iso_ts = dt.isoformat()
            return jsonify({
                "last_update": dt.strftime("%d/%m/%Y %H:%M:%S"),
                "timestamp": iso_ts,
                "formatted": dt.strftime("%d/%m/%Y %H:%M:%S"),
                "file": update_file,
            })
        else:
            now_iso = datetime.now().isoformat()
            return jsonify({
                "last_update": "File not found",
                "timestamp": now_iso,
                "formatted": datetime.fromisoformat(now_iso).strftime("%d/%m/%Y %H:%M:%S"),
            })
    except Exception as exc:
        return jsonify({"error": str(exc)}), 500


@app.route("/refresh_status")
def refresh_status():
    """Return the status of the last automatic or manual refresh."""
    return jsonify(last_refresh_status)


@app.route("/refresh_stock", methods=["POST"])
@admin_required
def refresh_stock():
    """Legacy alias for manual item stock refresh."""
    result = _refresh_stock_data()
    return jsonify(result), 200


# ----------------------------
# System panel routes (see the SYSTEM PANEL block above admin_required
# for the SYSTEM_ACCESS_TOKEN / device-pairing setup this all depends on)
# ----------------------------

@app.route("/admin/system/authorize-device")
@admin_required
def system_authorize_device():
    if not _system_panel_configured():
        return jsonify({
            "error": "System panel is not configured. Set SYSTEM_ACCESS_TOKEN in .env to enable it."
        }), 403

    token = request.args.get("token", "")
    if not token or token != Config.SYSTEM_ACCESS_TOKEN:
        return jsonify({"error": "Invalid access token."}), 403

    signed_value = _system_device_signer().sign(SYSTEM_DEVICE_COOKIE_VALUE).decode("utf-8")
    response = redirect(url_for("system_panel"))
    response.set_cookie(
        SYSTEM_DEVICE_COOKIE_NAME,
        signed_value,
        max_age=SYSTEM_DEVICE_COOKIE_MAX_AGE,
        httponly=True,
        secure=Config.SESSION_COOKIE_SECURE,
        samesite="Strict",
    )
    return response


@app.route("/admin/system")
@admin_required
@system_device_required
def system_panel():
    return render_template("system.html")


def _run_git_command(args, timeout=10):
    try:
        result = subprocess.run(
            ["git"] + list(args),
            cwd=BASE_DIR,
            capture_output=True,
            text=True,
            timeout=timeout,
            creationflags=subprocess.CREATE_NO_WINDOW,
        )
        return result.returncode, result.stdout.strip(), result.stderr.strip()
    except Exception as exc:
        return -1, "", str(exc)


@app.route("/admin/system/status")
@admin_required
@system_device_required
def system_status():
    _, local_commit, _ = _run_git_command(["log", "-1", "--format=%h %s"])
    _, local_commit_time, _ = _run_git_command(["log", "-1", "--format=%ci"])

    recent_code, recent_out, _ = _run_git_command(["log", "-10", "--format=%h|%s|%ci"])
    recent_commits = []
    if recent_code == 0:
        for line in recent_out.splitlines():
            parts = line.split("|", 2)
            if len(parts) == 3:
                recent_commits.append({"hash": parts[0], "message": parts[1], "time": parts[2]})

    remote_commit = None
    up_to_date = None
    offline = False
    try:
        fetch_code, _, fetch_err = _run_git_command(["fetch", "origin", "main"], timeout=15)
        if fetch_code != 0:
            raise RuntimeError(fetch_err or "git fetch failed")
        rev_code, rev_out, _ = _run_git_command(["rev-parse", "origin/main"])
        if rev_code == 0:
            remote_commit = rev_out
    except Exception:
        offline = True

    if remote_commit:
        head_code, head_out, _ = _run_git_command(["rev-parse", "HEAD"])
        if head_code == 0:
            up_to_date = (head_out == remote_commit)

    return jsonify({
        "local_commit": local_commit or None,
        "local_commit_time": local_commit_time or None,
        "recent_commits": recent_commits,
        "remote_commit": remote_commit,
        "up_to_date": up_to_date,
        "offline": offline,
    })


@app.route("/admin/system/logs")
@admin_required
@system_device_required
def system_logs():
    lines = request.args.get("lines", default=200, type=int) or 200
    lines = max(1, min(lines, 1000))

    log_path = Config.LOG_FILE
    if not os.path.isabs(log_path):
        log_path = os.path.join(BASE_DIR, log_path)

    if not os.path.exists(log_path):
        return Response("(log file not found)", mimetype="text/plain")

    try:
        with open(log_path, "r", encoding="utf-8", errors="replace") as handle:
            all_lines = handle.readlines()
    except OSError as exc:
        return Response(f"(log file could not be read right now: {exc})", mimetype="text/plain")

    return Response("".join(all_lines[-lines:]), mimetype="text/plain")


def _spawn_relaunch_helper():
    """Spawn relaunch_helper.py detached so the restart works even if
    launcher.pyw's watchdog isn't running (see relaunch_helper.py's
    docstring for why this can't just rely on that watchdog alone)."""
    python_w = os.path.join(BASE_DIR, ".venv", "Scripts", "pythonw.exe")
    python_exe = python_w if os.path.exists(python_w) else sys.executable
    helper_script = os.path.join(BASE_DIR, "relaunch_helper.py")
    try:
        subprocess.Popen(
            [python_exe, helper_script],
            cwd=BASE_DIR,
            creationflags=subprocess.CREATE_NO_WINDOW,
        )
        return True
    except Exception:
        logger.exception("Failed to spawn relaunch_helper.py")
        return False


def _trigger_self_restart(reason):
    def _do_restart():
        time.sleep(1.5)
        logger.info(reason)
        spawned = _spawn_relaunch_helper()
        if not spawned:
            logger.error("relaunch_helper.py failed to spawn; server will NOT come back automatically.")
        os._exit(0)

    threading.Thread(target=_do_restart, daemon=True).start()


@app.route("/admin/system/pull_and_restart", methods=["POST"])
@admin_required
@system_device_required
def system_pull_and_restart():
    code, out, err = _run_git_command(["pull", "origin", "main"], timeout=30)
    output = (out + ("\n" + err if err else "")).strip()

    if code != 0:
        return jsonify({"success": False, "restarted": False, "output": output or "git pull failed"})

    if "Already up to date" in out or "Already up-to-date" in out:
        return jsonify({"success": True, "restarted": False, "output": output or "Already up to date."})

    _trigger_self_restart("Restarting after code update via admin panel")
    return jsonify({"success": True, "restarted": True, "output": output})


@app.route("/admin/system/restart_app_only", methods=["POST"])
@admin_required
@system_device_required
def system_restart_app_only():
    _trigger_self_restart("Restarting via admin panel (no code pull)")
    return jsonify({"success": True, "restarted": True})


@app.route("/admin/system/download_backup")
@admin_required
@system_device_required
def system_download_backup():
    main_file_path = get_main_file_path()
    backup_targets = [
        (db.DB_PATH, "mappings.db"),
        (main_file_path, os.path.basename(main_file_path) if main_file_path else None),
        (CAR_FILE, os.path.basename(CAR_FILE)),
        (os.path.join(BASE_DIR, MAIN_HIERARCHY_CACHE_JSON), "main_hierarchy.json"),
        (os.path.join(BASE_DIR, CAR_MASTER_CACHE_JSON), "car_master.json"),
        (ITEM_STOCK_CACHE_JSON, "item stock list.auto.json"),
    ]

    buffer = BytesIO()
    with zipfile.ZipFile(buffer, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        for source_path, archive_name in backup_targets:
            if source_path and archive_name and os.path.exists(source_path):
                archive.write(source_path, arcname=archive_name)
    buffer.seek(0)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return send_file(
        buffer,
        mimetype="application/zip",
        as_attachment=True,
        download_name=f"backup_{timestamp}.zip",
    )


@app.route("/admin/system/find_duplicate_images")
@admin_required
@system_device_required
def system_find_duplicate_images():
    groups = db.find_duplicate_image_rows()
    total_extra_rows = sum(group["count"] - 1 for group in groups)
    return jsonify({
        "groups": groups,
        "group_count": len(groups),
        "total_extra_rows": total_extra_rows,
    })


def _tally_is_reachable(timeout=3):
    try:
        requests.get(TALLY_URL, timeout=timeout)
        return True
    except requests.RequestException:
        return False


@app.route("/admin/system/tally_status")
@admin_required
@system_device_required
def system_tally_status():
    reachable = _tally_is_reachable()
    instance_count = _check_multiple_tally_instances()

    warning = None
    if not reachable:
        warning = f"Tally is not reachable at {TALLY_URL}."
    elif instance_count > 1:
        warning = f"Multiple Tally instances detected ({instance_count} running). Close the duplicates and keep only one open."

    return jsonify({
        "reachable": reachable,
        "instance_count": instance_count,
        "warning": warning,
    })


AUTOSTART_REGISTRY_KEY = r"HKCU\Software\Microsoft\Windows\CurrentVersion\Run"
AUTOSTART_VALUE_NAME = "TallyStockViewer"


def _expected_autostart_command():
    pythonw_exe = os.path.join(BASE_DIR, ".venv", "Scripts", "pythonw.exe")
    launcher_pyw = os.path.join(BASE_DIR, "launcher.pyw")
    return f'"{pythonw_exe}" "{launcher_pyw}"'


@app.route("/admin/system/autostart_status")
@admin_required
@system_device_required
def system_autostart_status():
    expected = _expected_autostart_command()
    try:
        result = subprocess.run(
            ["reg", "query", AUTOSTART_REGISTRY_KEY, "/v", AUTOSTART_VALUE_NAME],
            capture_output=True,
            text=True,
            timeout=10,
            creationflags=subprocess.CREATE_NO_WINDOW,
        )
    except Exception as exc:
        return jsonify({"exists": False, "matches": False, "error": str(exc)})

    if result.returncode != 0:
        return jsonify({"exists": False, "matches": False, "current": None, "expected": expected})

    current = None
    for line in result.stdout.splitlines():
        stripped = line.strip()
        if stripped.startswith(AUTOSTART_VALUE_NAME):
            parts = stripped.split(None, 2)
            if len(parts) == 3:
                current = parts[2]
            break

    return jsonify({
        "exists": True,
        "matches": current == expected,
        "current": current,
        "expected": expected,
    })


# The same requests fetch_item_stock_flat() sends on every export cycle
# (A, B, C), plus the proposed lighter single-request replacement (D).
# Mirrors scripts/measure_tally.ps1 so browser and PowerShell runs are
# directly comparable. Read-only: export/collection requests only.
_TALLY_PERF_REQUESTS = [
    ("A. Item master collection (current)", "collection", '''<ENVELOPE>
    <HEADER>
        <VERSION>1</VERSION>
        <TALLYREQUEST>Export Data</TALLYREQUEST>
        <TYPE>Collection</TYPE>
        <ID>List of Stock Items</ID>
    </HEADER>
    <BODY>
        <DESC>
            <STATICVARIABLES>
                <SVEXPORTFORMAT>$$SysName:XML</SVEXPORTFORMAT>
            </STATICVARIABLES>
        </DESC>
    </BODY>
</ENVELOPE>'''),
    ("B. Stock Summary non-detailed (current)", "summary", '''<ENVELOPE>
    <HEADER>
        <VERSION>1</VERSION>
        <TALLYREQUEST>Export</TALLYREQUEST>
        <TYPE>Data</TYPE>
        <ID>Stock Summary</ID>
    </HEADER>
    <BODY>
        <DESC>
            <STATICVARIABLES>
                <SVEXPORTFORMAT>$$SysName:XML</SVEXPORTFORMAT>
            </STATICVARIABLES>
        </DESC>
    </BODY>
</ENVELOPE>'''),
    ("C. Stock Summary detailed+exploded (current)", "summary", '''<ENVELOPE>
    <HEADER>
        <VERSION>1</VERSION>
        <TALLYREQUEST>Export</TALLYREQUEST>
        <TYPE>Data</TYPE>
        <ID>Stock Summary</ID>
    </HEADER>
    <BODY>
        <DESC>
            <STATICVARIABLES>
                <SVEXPORTFORMAT>$$SysName:XML</SVEXPORTFORMAT>
                <SVSTOCKGROUP>Primary</SVSTOCKGROUP>
                <ISDETAILED>Yes</ISDETAILED>
                <EXPLODEFLAG>Yes</EXPLODEFLAG>
                <SVSHOWALLITEMS>Yes</SVSHOWALLITEMS>
                <SVSHOWZEROBALANCES>Yes</SVSHOWZEROBALANCES>
            </STATICVARIABLES>
        </DESC>
    </BODY>
</ENVELOPE>'''),
    ("D. Item collection with closing balance (candidate)", "collection", '''<ENVELOPE>
    <HEADER>
        <VERSION>1</VERSION>
        <TALLYREQUEST>Export Data</TALLYREQUEST>
        <TYPE>Collection</TYPE>
        <ID>Item Names With Closing</ID>
    </HEADER>
    <BODY>
        <DESC>
            <STATICVARIABLES>
                <SVEXPORTFORMAT>$$SysName:XML</SVEXPORTFORMAT>
            </STATICVARIABLES>
            <TDL>
                <TDLMESSAGE>
                    <COLLECTION NAME="Item Names With Closing" ISMODIFY="No">
                        <TYPE>StockItem</TYPE>
                        <FETCH>NAME, CLOSINGBALANCE</FETCH>
                    </COLLECTION>
                </TDLMESSAGE>
            </TDL>
        </DESC>
    </BODY>
</ENVELOPE>'''),
]


@app.route("/admin/system/tally_perf_test")
@admin_required
@system_device_required
def system_tally_perf_test():
    if not _tally_is_reachable():
        return jsonify({"ok": False, "error": f"Tally is not reachable at {TALLY_URL}."})
    if _check_multiple_tally_instances() > 1:
        return jsonify({"ok": False, "error": MULTIPLE_TALLY_MESSAGE})

    def _parse_collection(text):
        root = ET.fromstring(_sanitize_tally_xml(text))
        nodes = root.findall(".//COLLECTION/STOCKITEM")
        samples = []
        for node in nodes[:3]:
            name = node.attrib.get("NAME", "").strip()
            if not name:
                name_node = node.find(".//NAME")
                name = name_node.text.strip() if (name_node is not None and name_node.text) else ""
            bal_node = node.find(".//CLOSINGBALANCE")
            qty = bal_node.text.strip() if (bal_node is not None and bal_node.text) else ""
            samples.append({"name": name, "qty": qty})
        return len(nodes), samples

    def _parse_summary(text):
        root = ET.fromstring(_sanitize_tally_xml(text))
        names = root.findall(".//DSPACCNAME")
        stocks = root.findall(".//DSPSTKINFO")
        samples = []
        for name_node, stock_node in list(zip(names, stocks))[:3]:
            display_node = name_node.find("DSPDISPNAME")
            qty_node = stock_node.find(".//DSPCLQTY")
            name = display_node.text.strip() if (display_node is not None and display_node.text) else ""
            qty = qty_node.text.strip() if (qty_node is not None and qty_node.text) else ""
            samples.append({"name": name, "qty": qty})
        return len(names), samples

    parsers = {"collection": _parse_collection, "summary": _parse_summary}

    results = []
    for label, kind, xml_req in _TALLY_PERF_REQUESTS:
        start = time.perf_counter()
        try:
            text = _post_tally_with_retry(xml_req)
            elapsed = time.perf_counter() - start
            if "<LINEERROR>" in text or "<RESPONSE>Error" in text or "Unknown Request" in text:
                results.append({"label": label, "time_seconds": round(elapsed, 2),
                                "status": f"Tally error: {text[:200]}", "row_count": None, "samples": []})
                continue
            try:
                row_count, samples = parsers[kind](text)
                results.append({"label": label, "time_seconds": round(elapsed, 2),
                                "status": "ok", "row_count": row_count, "samples": samples})
            except Exception as parse_exc:
                results.append({"label": label, "time_seconds": round(elapsed, 2),
                                "status": f"parse error: {parse_exc}", "row_count": None, "samples": []})
        except Exception as exc:
            elapsed = time.perf_counter() - start
            results.append({"label": label, "time_seconds": round(elapsed, 2),
                            "status": str(exc), "row_count": None, "samples": []})

    return jsonify({"ok": True, "results": results})


@app.route("/admin/system/env_summary")
@admin_required
@system_device_required
def system_env_summary():
    # Deliberately excluded from this summary and never sent to the
    # browser: FLASK_SECRET_KEY, SYSTEM_ACCESS_TOKEN.
    return jsonify({
        "TALLY_URL": Config.TALLY_URL,
        "DB_PATH": db.DB_PATH,
        "LOG_FILE": Config.LOG_FILE,
        "IMAGE_SCAN_ROOT": IMAGE_SCAN_ROOT,
        "SESSION_COOKIE_SECURE": Config.SESSION_COOKIE_SECURE,
        "FLASK_DEBUG": Config.DEBUG,
    })


DISK_USAGE_CACHE = {"timestamp": 0.0, "image_folder_size_mb": None}


def _get_image_folder_size_mb():
    now = time.time()
    if DISK_USAGE_CACHE["image_folder_size_mb"] is not None and now - DISK_USAGE_CACHE["timestamp"] < 60:
        return DISK_USAGE_CACHE["image_folder_size_mb"]

    total_bytes = 0
    for root, _dirs, files in os.walk(IMAGE_SCAN_ROOT):
        for filename in files:
            try:
                total_bytes += os.path.getsize(os.path.join(root, filename))
            except OSError:
                continue

    size_mb = round(total_bytes / (1024 * 1024), 2)
    DISK_USAGE_CACHE["timestamp"] = now
    DISK_USAGE_CACHE["image_folder_size_mb"] = size_mb
    return size_mb


@app.route("/admin/system/disk_usage")
@admin_required
@system_device_required
def system_disk_usage():
    db_size_mb = round(os.path.getsize(db.DB_PATH) / (1024 * 1024), 2) if os.path.exists(db.DB_PATH) else 0.0
    free_bytes = shutil.disk_usage(BASE_DIR).free
    return jsonify({
        "mappings_db_size_mb": db_size_mb,
        "image_folder_size_mb": _get_image_folder_size_mb(),
        "free_disk_space_gb": round(free_bytes / (1024 ** 3), 2),
    })


def _format_uptime(seconds):
    seconds = int(seconds)
    days, remainder = divmod(seconds, 86400)
    hours, remainder = divmod(remainder, 3600)
    minutes, _ = divmod(remainder, 60)
    parts = []
    if days:
        parts.append(f"{days}d")
    if days or hours:
        parts.append(f"{hours}h")
    parts.append(f"{minutes}m")
    return " ".join(parts)


@app.route("/admin/system/uptime")
@admin_required
@system_device_required
def system_uptime():
    elapsed = time.time() - START_TIME
    return jsonify({
        "uptime_seconds": round(elapsed, 1),
        "uptime_human": _format_uptime(elapsed),
    })


# ----------------------------
# System panel: category reorder + abbreviation override (Part 6, advanced
# controls tucked behind the same admin + device-pairing gate as every other
# System panel feature -- the everyday add/rename/delete flow lives instead
# in index.html's lightweight Category Settings panel, see Part 5).
# ----------------------------

@app.route("/admin/system/categories")
@admin_required
@system_device_required
def system_list_categories():
    return jsonify({"categories": db.get_all_categories()})


@app.route("/admin/system/categories/move", methods=["POST"])
@admin_required
@system_device_required
def system_move_category():
    payload = request.get_json(silent=True) or {}
    name = str(payload.get("name") or "").strip()
    direction = str(payload.get("direction") or "").strip().lower()

    if not db.category_exists(name):
        return jsonify({"error": f"category '{name}' does not exist"}), 404
    if direction not in ("up", "down"):
        return jsonify({"error": "direction must be 'up' or 'down'"}), 400

    try:
        moved = db.move_category(name, direction)
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400

    return jsonify({"moved": moved, "categories": db.get_all_categories()})


@app.route("/admin/system/categories/abbreviation", methods=["POST"])
@admin_required
@system_device_required
def system_update_category_abbreviation():
    payload = request.get_json(silent=True) or {}
    name = str(payload.get("name") or "").strip()
    abbreviation = payload.get("abbreviation")

    if not db.category_exists(name):
        return jsonify({"error": f"category '{name}' does not exist"}), 404

    try:
        new_abbreviation = db.update_category_abbreviation(name, abbreviation)
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400

    logger.info("Category '%s' abbreviation set to '%s' by user_id=%s", name, new_abbreviation, _current_user_id())

    # The burned share-image badge now uses the category's full NAME, not
    # the abbreviation (see _get_category_info_for_image()), so this
    # regeneration no longer changes the badge's visible text -- kept as-is
    # since it's harmless (same text re-burned) and this route is Category
    # Settings territory, out of scope for the badge-source fix above. Only
    # the on-thumbnail .category-ribbon (client-side) actually reflects an
    # abbreviation edit, and it updates instantly via refreshCategoryData()
    # with no server-side regeneration needed.
    affected_names = _stock_item_names_for_category(name)
    if affected_names:
        _regenerate_badges_for_stock_items(affected_names, category_name=name)

    return jsonify({"name": name, "abbreviation": new_abbreviation, "affected_count": len(affected_names)})


set_search_dependencies(
    get_car_models=lambda: list(CAR_GROUPS),
    get_car_folders=db.get_image_folders,
    get_stock_items_for_car_from_hierarchy=get_stock_items_for_training_from_hierarchy,
    get_stock_items_for_car=get_stock_items_for_training,
)


# ----------------------------
# Startup
# ----------------------------
if __name__ == "__main__":
    # Dual-file mode:
    # - main.xlsx/main.xls is manual and holds hierarchy
    # - item stock list.xls is auto-exported and holds quantities
    
    print("=" * 60)
    print("Tally Stock Viewer - Hierarchical Parser")
    print("=" * 60)
    
    main_file = get_main_file_path()
    if not os.path.exists(main_file):
        print(f"\nWARNING: Main file not found. Tried: {', '.join(MAIN_FILE_CANDIDATES)}")
        print("\nmain.xlsx/main.xls is required for parent-child hierarchy.")
        print("Auto-export updates only item stock list.xlsx (quantities).")
        print("Please manually export hierarchy once and save as data/main.xlsx")
    else:
        print(f"Using main hierarchy file: {main_file}")
        print("Starting background data load and image scan...")

    start_background_startup_tasks()
    
    print("\nStarting Flask server on http://localhost:5000")
    print("Press Ctrl+C to stop\n")
    app.run(debug=app.config["DEBUG"])
